################################################################################################
################################################################################################
#################### Frequency Analysis with Patched/Recorded Rainfall Data ####################
################################################################################################
################################################################################################

#Last Updated: June 28, 2024

#This Python script generates a comprehensive Excel report encompassing various rainfall data
#analyses and visualisations.

#The report includes the original rainfall data, which is graphically represented by daily
#rainfall and classified in 5mm/d increments (e.g., 6.6179mm is classified as 5mm). It features
#a One Day Analysis section showing occurrence, exceedance, rate, and probability, accompanied
#by a Rainfall Occurrence graph and a Rainfall Probability graph.
#The One Day Probability Data Frame presents the return period and corresponding one-day
#probability, alongside a Probability of Rainfall Event graph.
#The Gumbel Estimate section details the maximum one-day rainfall, observations, rank by
#maximum one-day rainfall, exceedance, non-exceedance, return period, return variables,
#and the Gumbel Estimate, with graphs illustrating maximum one-day rainfall in a year and
#the Gumbel Estimate.
#The Gumbel Estimate and Full Series section includes return period, return variable,
#Gumbel Estimate, and Full Series data, along with a graph comparing
#Gumbel Estimate (mm/day) vs. Full Series (mm/day).
#The Residual Mass section displays yearly total rainfall, accumulated rainfall, average
#yearly rainfall, total rainfall minus average rainfall, and accumulated rainfall minus
#average rainfall, with accompanying graphs for total yearly rainfall, mass plot, and
#residual mass.
#The Monthly Rainfall (Box and Whisker) section lists monthly rainfall data and total
#yearly rainfall (mm), supplemented by a box and whisker graph and a total monthly
#rainfall graph.
#The Log Pearson (III) section presents frequency factors for gamma and Log-Pearson
#Type III distributions, return period, Wt value, log Qt, Pearson Estimate (mm/day),
#and Full Series (mm/day). - The GEV and MEV section provides csi, sigma, mu, and
#starting guesses for MEV (x0). The LP (III) Estimate section includes observed max
#rainfall (mm/day), rank, log observed max, statistical calculations, return period,
#exceedance probability, non-exceedance probability, Wt, and LP (III) Estimate (mm/day),
#along with a Log Pearson (III) Estimate graph.
#The LP Est, GEV, and MEV section compares return period, Full Series (mm/day), Gumbel
#Estimate (mm/day), Log Pearson (III) Estimate (mm/day), GEV Estimate (mm/day), and MEV
#Estimate (mm/day) with a comparative graph of all estimates.

#*********************************************************************************************
#*********************************************************************************************

import numpy as np
import pandas as pd
import math
import statistics
import openpyxl
from openpyxl import load_workbook
from openpyxl.chart import Reference, BarChart, LineChart, series
import time
from openpyxl.styles import NamedStyle, PatternFill, Border, Side, Alignment, Protection, Font
from openpyxl.chart import ScatterChart, Reference, Series
from datetime import datetime
import os
import plotly.graph_objects as go
import plotly.offline as pyo
from IPython.display import Image
import shutil
import mevpy as mev
import warnings
warnings.simplefilter('ignore') # filter some warning messages
import dataframe_image as dfi
import tkinter as tk
from tkinter import filedialog, messagebox
from tkinter import font as tkfont

################################################################################################
###################################### CREATE DATAFRAMES #######################################
################################################################################################
def continue_rainfall_analysis(orig_rain_data_path, external_path, python_path, downloads_path):
    #2. Define and Load User Input
    #2.1 Create 'df' Dataframe
    df = pd.read_excel(orig_rain_data_path)

    #***********************************************************************************************
    #2.2 Change NaN values to 0 inorder to do Calculations
    df.fillna(0, inplace=True)

    #***********************************************************************************************
    #2.3 Insert 'Patched Rainfall (mm/d)' Column
    patched_rainfall = []
    for i in range(len(df)):
        if df['Patched/Recorded'].iloc[i] == 'Patched':
            patched_rainfall.append(df['Rainfall (mm/d)'].iloc[i])
        else:
            patched_rainfall.append('')
    df.insert(3, 'Patched Rainfall (mm/d)', patched_rainfall)

    #***********************************************************************************************
    #2.4 Insert 'Recorded Rainfall (mm/d)' Column
    recorded_rainfall = []
    for i in range(len(df)):
        if df['Patched/Recorded'].iloc[i] == 'Recorded':
            recorded_rainfall.append(df['Rainfall (mm/d)'].iloc[i])
        else:
            recorded_rainfall.append(' ')
    df.insert(4, 'Recorded Rainfall (mm/d)', recorded_rainfall)

    ################################################################################################
    ################################## 3. Original Rainfall Data ###################################
    ################################################################################################
    #3.1 Create 'og_rainfall' DataFrame
    og_rainfall = df

    #**************************************************************************************************
    #3.2 Insert 'Class (5mm/d)' Column
    og_rainfall.insert(1, 'Class (5mm/d)', og_rainfall['Rainfall (mm/d)'])

    #**************************************************************************************************
    #3.3 Round 'Class (5mm/d)' Column down to the nearest 5
    og_rainfall['Class (5mm/d)'] = ((og_rainfall['Class (5mm/d)'])/5)
    og_rainfall['Class (5mm/d)'] = ((og_rainfall['Class (5mm/d)']).apply(np.floor))
    og_rainfall['Class (5mm/d)'] = ((og_rainfall['Class (5mm/d)'])*5)

    #**************************************************************************************************
    #3.4 Display 'og_rainfall' Dataframe
    df['Date']= pd.to_datetime(df['Date'])

    og_rainfall['Date']= pd.to_datetime(og_rainfall['Date'])

    ################################################################################################
    ############################### 4. Create 'oneday_df' Dataframe ################################
    ################################################################################################
    #4.1 Create 'oneday_df' DataFrame
    oneday_df = pd.DataFrame()

    #**************************************************************************************************
    #4.2 Insert 'Class Interval (mm)' Column
    class_intervals = []
    for i in range(int(((og_rainfall['Class (5mm/d)'].max()) + 10)/5)):
            class_intervals.append(i*5)

    oneday_df.insert(0, 'Class Interval (mm)', class_intervals)

    #**************************************************************************************************
    #4.3 Insert 'n_occ_1-Day' Column
    #4.3.1 Count all Class Occurences
    count_occ = og_rainfall.groupby('Class (5mm/d)')['Class (5mm/d)'].count()

    #4.3.2 Create 'not_final_count_occ_df' Dataframe
    not_final_count_occ_df = pd.DataFrame()

    #4.3.2.1 Insert 'Class Interval (mm)' Column
    not_final_count_occ_class = list(count_occ.index.values)
    not_final_count_occ_df.insert(0, 'Class Interval (mm)', not_final_count_occ_class)

    #4.3.2.2 Insert 'Occurences' Column
    not_final_count_occ_values = list(count_occ)
    not_final_count_occ_df.insert(1, 'Occurences', not_final_count_occ_values)

    #4.3.3 Insert 'n_occ_1-Day' Column
    final_count_occ = []

    for i in range(len(oneday_df)):
        if oneday_df['Class Interval (mm)'].iloc[i] in not_final_count_occ_df['Class Interval (mm)'].values:
            get_index = oneday_df['Class Interval (mm)'].iloc[i]
            get_occ = not_final_count_occ_df['Occurences'][not_final_count_occ_df['Class Interval (mm)'] == get_index]
            final_count_occ.append(int(get_occ))
        else:
            final_count_occ.append(0)

    oneday_df.insert(1, 'n_occ_1-Day', final_count_occ)

    #**************************************************************************************************
    #4.4 Insert 'n_ex_1-Day' Column
    final_count_exceeds = []
    for i in range(len(oneday_df)):
        if i != len(oneday_df):
            n = int(len(oneday_df))-1
            exceeds_count = oneday_df['n_occ_1-Day'][i:n].sum()
            final_count_exceeds.append(exceeds_count)
        else:
            final_count_exceeds.append(0)

    oneday_df.insert(2, 'n_ex_1-Day', final_count_exceeds)

    #**************************************************************************************************
    #4.5 Insert 'Rate(1/Year)' Column
    #4.5.1 Define amount of Years of Rainfall Data
    start_year = (og_rainfall['Date'].min()).year
    end_year = (og_rainfall['Date'].max()).year

    n_year = end_year - start_year

    #4.5.2 Insert 'Rate(1/Year)' Column
    rate_year = []
    for i in range(len(oneday_df)):
        rate_year.append(oneday_df['n_ex_1-Day'].iloc[i]/n_year)

    oneday_df.insert(3, 'Rate(1/Year)', rate_year)

    #**************************************************************************************************
    #4.6 Insert 'T(Year)' Column
    t_year = []
    for i in range(len(oneday_df)):
        t_year.append(1/(oneday_df['Rate(1/Year)']).iloc[i])

    oneday_df.insert(4, 'T(Year)', t_year)

    #**************************************************************************************************
    #4.7 Insert 'P(1/Year)' Column
    p_year = []
    for i in range(len(oneday_df)):
        p_formula = 1-(math.exp(float((-(oneday_df['Rate(1/Year)'].iloc[i]))*float(oneday_df['T(Year)'].iloc[i]))))
        p_year.append(float(p_formula))

    oneday_df.insert(5, 'P(1/Year)', p_year)

    ################################################################################################
    ################################### 5. Create 'DDF' Dataframe ##################################
    ################################################################################################
    #5.1 Create 'DDF' Dataframe
    ddf_df = pd.DataFrame()

    #**************************************************************************************************
    #5.2 Insert 'Return Period T (Year)' Column
    return_period = [0.5, 1, 1.2, 1.5, 2, 5, 10, 20, 35, 50, 100, 200, 500, 1000]
    ddf_df.insert(0, 'Return Period T (Year)', return_period)

    #**************************************************************************************************
    #5.3 Insert '1-Day P Sum (mm)' Column

    #'1-Day P Sum' = ((((c2-c1) / ((np.log10(t2)) - (np.log10(t1)))) * ((np.log10(rval)) - (np.log10(t1)))) + c1)
    #t1 = the value in the 'T (Year)' column closest to the 'Return Period T (year)' AND is less than 'Return Period T (year)'
    #t2 = the value in the 'T (Year)' column closest to the 'Return Period T (year)' AND is more than 'Return Period T (year)'
    #c1 = class interval that contains T1
    #c2 = class interval that contains T2
    #rval = Value in Return Period T (year)

    #Filter out infinity values
    filtered_values = oneday_df['T(Year)'][~np.isinf(oneday_df['T(Year)'])]
    #Get the maximum value
    max_value = filtered_values.max()

    psum_results = []
    for i in range(len(ddf_df)):
        if ddf_df['Return Period T (Year)'].iloc[i] <= max_value:
            rval = ddf_df['Return Period T (Year)'][i]
            for j in range(len(oneday_df)):
                if (oneday_df['T(Year)'][j] <= rval) and (oneday_df['T(Year)'][j+1] >= rval):
                    # Capture time(year)
                    t1 = oneday_df['T(Year)'][j]
                    t2 = oneday_df['T(Year)'][j+1]
                    # Capture class
                    c1 = oneday_df['Class Interval (mm)'][j]
                    c2 = oneday_df['Class Interval (mm)'][j+1]
                    break
            psum = ((((c2-c1) / ((np.log10(t2)) - (np.log10(t1)))) * ((np.log10(rval)) - (np.log10(t1)))) + c1)
            psum_results.append(psum)
        else:
            psum_results.append('')

    ddf_df.insert(1, '1-Day P Sum (mm)', psum_results)

    ################################################################################################
    ################################ 6. Create 'year_df' Dataframe #################################
    ################################################################################################
    #6.1 Create 'year_df' Dataframe
    year_df = pd.DataFrame()

    #**************************************************************************************************
    #6.2 Insert 'Year' Column
    years = list(pd.DatetimeIndex(og_rainfall['Date']).year)
    year_df.insert(0, 'Year', years)

    #**************************************************************************************************
    #6.3 Insert 'Rainfall (mm/d)' Column
    rainfall_data = []
    for i in range(len(year_df)):
        rainfall_data.append(og_rainfall['Rainfall (mm/d)'].iloc[i])

    year_df.insert(1, 'Rainfall (mm/d)', rainfall_data)

    ################################################################################################
    ############################## 7. Create 'max1day_df' Dataframe ################################
    ################################################################################################
    #7.1 Create 'max1day_df' Dataframe
    max1day_df = pd.DataFrame()

    #**************************************************************************************************
    #7.2 Insert 'Year' Column
    year_list = []
    for i in range(n_year+1):
        years_append = start_year + i
        year_list.append(years_append)
    max1day_df.insert(0, 'Year', year_list)

    #**************************************************************************************************
    #7.3 Insert 'Max 1-Day Rainfall (mm)' Column
    max_1day_rain = year_df.groupby('Year')['Rainfall (mm/d)'].max()

    max_rain = []
    for i in range(len(max1day_df)):
        max_rain.append(max_1day_rain.iloc[i])

    max1day_df.insert(1, 'Max 1-Day Rainfall (mm)', max_rain)

    ################################################################################################
    ############################## 8. Create 'gumbelest_df' Dataframe ##############################
    ################################################################################################
    #8.1 Create 'gumbelest_df' Dataframe
    gumbelest_df = pd.DataFrame()

    #**************************************************************************************************
    #8.2 Insert 'Year' Column
    gumbelest_df.insert(0, 'Year', year_list)

    #**************************************************************************************************
    #8.3 Insert 'Max 1-Day Rainfall (mm)' Column
    gumbelest_df.insert(1, 'Max 1-Day Rainfall (mm)', max_rain)

    #**************************************************************************************************
    #8.4 Insert 'N (#obs)' Column
    observation_n = n_year + 1

    n_obs = []
    for i in range(len(gumbelest_df)):
        n_obs.append(observation_n)

    gumbelest_df.insert(2, 'N (#obs)', n_obs)

    #**************************************************************************************************
    #8.5 Insert 'i (rank)' Column
    i_rank = []
    for i in range(len(gumbelest_df)):
        i_rank.append(i+1)

    gumbelest_df = gumbelest_df.sort_values(['Max 1-Day Rainfall (mm)'], ascending = [False])

    gumbelest_df.insert(3, 'i (Rank)', i_rank)

    #**************************************************************************************************
    #8.6 Insert 'P (Exceedance)' Column
    #P=i/(N+1)
    p_exceed = []
    for i in range(len(gumbelest_df)):
        p_exceed.append((gumbelest_df['i (Rank)'].iloc[i])/((gumbelest_df['N (#obs)'].iloc[i])+1))

    gumbelest_df.insert(4, 'P (Exceedance)', p_exceed)

    #**************************************************************************************************
    #8.7 Insert 'Q (Non-Exceedance)' Column
    #Q = 1 - P
    q_non = []
    for i in range(len(gumbelest_df)):
        q_non.append(1 - (gumbelest_df['P (Exceedance)'].iloc[i]))

    gumbelest_df.insert(5, 'Q (Non-Exc)', q_non)

    #**************************************************************************************************
    #8.8 Insert 'T_a(Return Period)' Column
    #T_a = 1/P
    t_a = []
    for i in range(len(gumbelest_df)):
        t_a.append(1 / (gumbelest_df['P (Exceedance)'].iloc[i]))

    gumbelest_df.insert(6, 'T_a (Return Period)', t_a)

    #**************************************************************************************************
    #8.9 Insert 'Y (Reduced Var)' Column
    #-ln(-ln(Q))
    y_redvar = []
    for i in range(len(gumbelest_df)):
        y_redvar.append(-np.log(-np.log(gumbelest_df['Q (Non-Exc)'].iloc[i])))

    gumbelest_df.insert(7, 'Y (Return Var)', y_redvar)

    #**************************************************************************************************
    #8.10 Insert 'Gumbel Estimate (mm/day)' Column
    #(Sigma*[Y (Reduced Var)]) + mu

    #8.10.1 Define Variables
    #s_y = Std Dev of Y (Reduced Var)
    s_y = np.std(y_redvar)

    #s_R = Std Dev of Max 1-Day Rainfall (mm)
    s_R = np.std(max1day_df['Max 1-Day Rainfall (mm)'])

    #R_max_gem = Average of Max 1-Day Rainfall (mm)
    R_max_gem = np.average(max1day_df['Max 1-Day Rainfall (mm)'])

    #y_gem = Average of Y (Reduced Var)
    y_gem = np.average(y_redvar)

    #s_y = Std Dev of Y (Reduced Var)
    s_y = np.std(y_redvar)

    #Sigma = s_R / s_y
    sigma = s_R / s_y

    #mu = R_max_gem-((s_R)*(y_gem/s_y))
    mu = R_max_gem-((s_R)*(y_gem / s_y))

    #8.10.2 (Sigma*[Y (Reduced Var)]) + mu
    gumbel_est = []
    for i in range(len(gumbelest_df)):
        gumbel_est.append(sigma*(y_redvar[i])+mu)

    gumbelest_df.insert(8, 'Gumbel Estimate (mm/d)', gumbel_est)

    ################################################################################################
    ############################## 9. Create 'fullgumbel_df' Dataframe #############################
    ################################################################################################
    #9.1 Create 'fullgumbel_df' Dataframe
    fullgumbel_df = pd.DataFrame()

    #**************************************************************************************************
    #9.2 Insert 'Return Period T (Year)' Column
    fullgumbel_df.insert(0, 'Return Period T (Year)', return_period)

    #**************************************************************************************************
    #9.3 Insert 'T_a (Gumbel)' Column
    #(1)/(1-e^(-1/['Return Period T (Year)']))
    t_a_gumbel = []
    for i in range(len(fullgumbel_df)):
        t_a_gumbel.append(1 / (1 - (math.exp(float(-1 / (fullgumbel_df['Return Period T (Year)'].iloc[i]))))))

    fullgumbel_df.insert(1, 'T_a (Gumbel)', t_a_gumbel)

    #**************************************************************************************************
    #9.4 Insert 'y (Gumbel)' Column
    #-ln(-ln(1-(1/['T_a (Gumbel)'])))
    y_gumbel = []
    for i in range(len(fullgumbel_df)):
        y_gumbel.append(-np.log(-np.log(1 - (1 / (fullgumbel_df['T_a (Gumbel)'].iloc[i])))))

    fullgumbel_df.insert(2, 'Y (Gumbel)', y_gumbel)

    #**************************************************************************************************
    #9.5 Insert 'Gumbel Estimate (mm/day)' Column
    #sigma* y_gumbel[] + mu
    est_gumbel = []
    for i in range(len(fullgumbel_df)):
        est_gumbel.append(sigma * (fullgumbel_df['Y (Gumbel)'].iloc[i]) + mu)

    fullgumbel_df.insert(3, 'Gumbel Estimate (mm/d)', est_gumbel)

    #**************************************************************************************************
    #9.6 Insert 'Full Series (mm/day)' Column
    #Create Full Series (mm/day) Column
    fullgumbel_df.insert(4, 'Full Series (mm/d)', ddf_df['1-Day P Sum (mm)'])

    #**************************************************************************************************
    #9.7 Replace 'n/a' values with ' '
    fullgumbel_df = fullgumbel_df.replace({'n/a': ''})

    ################################################################################################
    ############################### 10. Create 'resmass_df' Dataframe ##############################
    ################################################################################################
    #10.1 Create 'resmass_df' Dataframe
    resmass_df = pd.DataFrame()

    #**************************************************************************************************
    #10.2 Insert 'Year' Column
    resmass_df.insert(0, 'Year', year_list)

    #**************************************************************************************************
    #10.3 Insert 'Total Rainfall in Year' Column
    sum_rain = year_df.groupby('Year')['Rainfall (mm/d)'].sum()

    yearly_sum_rain = []
    for i in range(len(resmass_df)):
        yearly_sum_rain.append(sum_rain.iloc[i])

    resmass_df.insert(1, 'Total Rainfall in Year', yearly_sum_rain)

    #**************************************************************************************************
    #10.4 Insert 'Accumulated Rainfall' Column
    acc_rain = []
    for i in range(len(resmass_df)):
        if i == 0:
            acc_rain.append(resmass_df['Total Rainfall in Year'].iloc[0])
        else:
            acc_rain.append(resmass_df['Total Rainfall in Year'].head(i+1).sum())

    resmass_df.insert(2, 'Accumulated Rainfall', acc_rain)

    #**************************************************************************************************
    #10.5 Insert 'Average Yearly Rainfall' Column
    x_avg = np.average(resmass_df['Total Rainfall in Year'])

    x_average = []
    for i in range(len(resmass_df)):
        x_average.append(x_avg)

    resmass_df.insert(3, 'Average Yearly Rainfall', x_average)

    #**************************************************************************************************
    #10.6 Insert 'Total Rainfall in Year - Average Rainfall' Column
    x_xavg = []
    for i in range(len(resmass_df)):
        x_xavg.append(resmass_df['Total Rainfall in Year'].iloc[i] - x_avg)

    resmass_df.insert(4, 'Total Rainfall in Year - Average Rainfall', x_xavg)

    #**************************************************************************************************
    #10.7 Insert 'Accumulated Rainfall in Year - Average Rainfall'  Column
    accx_xavg = []
    for i in range(len(resmass_df)):
        if i == 0:
            accx_xavg.append(resmass_df['Total Rainfall in Year - Average Rainfall'].iloc[0])
        else:
            accx_xavg.append(accx_xavg[i-1] + resmass_df['Total Rainfall in Year - Average Rainfall'].iloc[i])

    resmass_df.insert(5, 'Accumulated Rainfall in Year - Average Rainfall', accx_xavg)

    ################################################################################################
    ############################# 11. Create 'boxwhisker_df' Dataframe #############################
    ################################################################################################
    #11.1 Create 'modifieddata_df' Dataframe
    modifieddata_df = pd.DataFrame()

    #**************************************************************************************************
    #11.2 Insert 'Date' Column
    og_rainfall['Date'] = og_rainfall['Date'].astype(str)

    yyyy_mm_date = []
    for i in range(len(og_rainfall)):
        yyyy_mm_date.append(og_rainfall['Date'].iloc[i][:7])

    modifieddata_df.insert(0, 'Date', yyyy_mm_date)

    #**************************************************************************************************
    #11.3 Insert 'Rainfall (mm/d)' Column
    modifieddata_df.insert(1, 'Rainfall (mm/d)', og_rainfall['Rainfall (mm/d)'])

    #**************************************************************************************************
    #11.4 Create 'boxwhisker_df' Dataframe
    #11.4.1 Create 'boxwhisker_df' Dataframe
    boxwhisker_df = pd.DataFrame()

    #11.4.2 Insert 'Year and Month' Column
    rain_sum_month = modifieddata_df.groupby('Date')['Rainfall (mm/d)'].sum()

    monthly_rain = []
    for i in range(len(rain_sum_month)):
        monthly_rain.append(rain_sum_month[i])

    monthly_dates = list(rain_sum_month.index.values)

    boxwhisker_df.insert(0, 'Year and Month', monthly_dates)

    #11.4.3 Insert 'Monthly Total Rainfall (mm)' Column
    boxwhisker_df.insert(1, 'Monthly Total Rainfall (mm)', monthly_rain)

    ################################################################################################
    ############################## 12. Create 'monthly_df' Dataframe ###############################
    ################################################################################################
    #12.1 Create 'monthly_df' Dataframe
    monthly_df = pd.DataFrame()

    #**************************************************************************************************
    #12.2 Insert 'Year' Column
    monthly_df.insert(0, 'Year', year_list)

    #**************************************************************************************************
    #12.3 Create 'monthly_sum_df' Dataframe
    #12.3.1 Create 'monthly_sum_df' Dataframe
    monthly_sum_df = pd.DataFrame()

    #12.3.2 Insert 'Year and Month' Column
    _0 = '-0'
    _hyphen = '-'

    year_month_full_series = []
    for i in range(len(year_list)):
        for j in range(13):
            if j != 0:
                if j < 10:
                    year_month_full_series.append(str(year_list[i]) + _0 + str(j))
                else:
                    year_month_full_series.append(str(year_list[i]) + _hyphen + str(j))

    monthly_sum_df.insert(0, 'Year and Month', year_month_full_series)

    #12.3.3 Insert 'Monthly Rainfall' Column
    rain_month_full_series = []
    for i in range(len(monthly_sum_df)):
        if monthly_sum_df['Year and Month'].iloc[i] in boxwhisker_df['Year and Month'].values:
            get_year_month = monthly_sum_df['Year and Month'].iloc[i]
            get_rainfall = boxwhisker_df['Monthly Total Rainfall (mm)'][boxwhisker_df['Year and Month'] == get_year_month]
            rain_month_full_series.append(float(get_rainfall))
        else:
            rain_month_full_series.append(0)

    monthly_sum_df.insert(1, 'Monthly Rainfall', rain_month_full_series)

    #**************************************************************************************************
    #12.4 Insert 'January' - 'December' Columns
    #12.4.1 Insert 'January' Column
    january = []
    for i in range(len(monthly_sum_df)):
        if monthly_sum_df['Year and Month'].iloc[i][5:]  == '01':
            m_date = monthly_sum_df['Year and Month'].iloc[i]
            month_rain = (monthly_sum_df['Monthly Rainfall'][monthly_sum_df['Year and Month'] == m_date])
            month_sum_rain = float(month_rain)
            january.append(month_sum_rain)

    monthly_df.insert(1, 'January', january)

    #12.4.2 Insert 'February' Column
    february = []
    for i in range(len(monthly_sum_df)):
        if monthly_sum_df['Year and Month'].iloc[i][5:]  == '02':
            m_date = monthly_sum_df['Year and Month'].iloc[i]
            month_rain = (monthly_sum_df['Monthly Rainfall'][monthly_sum_df['Year and Month'] == m_date])
            month_sum_rain = float(month_rain)
            february.append(month_sum_rain)

    monthly_df.insert(2, 'February', february)

    #12.4.3 Insert 'March' Column
    march = []
    for i in range(len(monthly_sum_df)):
        if monthly_sum_df['Year and Month'].iloc[i][5:]  == '03':
            m_date = monthly_sum_df['Year and Month'].iloc[i]
            month_rain = (monthly_sum_df['Monthly Rainfall'][monthly_sum_df['Year and Month'] == m_date])
            month_sum_rain = float(month_rain)
            march.append(month_sum_rain)

    monthly_df.insert(3, 'March', march)

    #12.4.4 Insert 'April' Column
    april = []
    for i in range(len(monthly_sum_df)):
        if monthly_sum_df['Year and Month'].iloc[i][5:]  == '04':
            m_date = monthly_sum_df['Year and Month'].iloc[i]
            month_rain = (monthly_sum_df['Monthly Rainfall'][monthly_sum_df['Year and Month'] == m_date])
            month_sum_rain = float(month_rain)
            april.append(month_sum_rain)

    monthly_df.insert(4, 'April', april)

    #12.4.5 Insert 'May' Column
    may = []
    for i in range(len(monthly_sum_df)):
        if monthly_sum_df['Year and Month'].iloc[i][5:]  == '05':
            m_date = monthly_sum_df['Year and Month'].iloc[i]
            month_rain = (monthly_sum_df['Monthly Rainfall'][monthly_sum_df['Year and Month'] == m_date])
            month_sum_rain = float(month_rain)
            may.append(month_sum_rain)

    monthly_df.insert(5, 'May', may)

    #12.4.6 Insert 'June' Column
    june = []
    for i in range(len(monthly_sum_df)):
        if monthly_sum_df['Year and Month'].iloc[i][5:]  == '06':
            m_date = monthly_sum_df['Year and Month'].iloc[i]
            month_rain = (monthly_sum_df['Monthly Rainfall'][monthly_sum_df['Year and Month'] == m_date])
            month_sum_rain = float(month_rain)
            june.append(month_sum_rain)

    monthly_df.insert(6, 'June', june)

    #12.4.7 Insert 'July' Column
    july = []
    for i in range(len(monthly_sum_df)):
        if monthly_sum_df['Year and Month'].iloc[i][5:]  == '07':
            m_date = monthly_sum_df['Year and Month'].iloc[i]
            month_rain = (monthly_sum_df['Monthly Rainfall'][monthly_sum_df['Year and Month'] == m_date])
            month_sum_rain = float(month_rain)
            july.append(month_sum_rain)

    monthly_df.insert(7, 'July', july)

    #12.4.8 Insert 'August' Column
    august = []
    for i in range(len(monthly_sum_df)):
        if monthly_sum_df['Year and Month'].iloc[i][5:]  == '08':
            m_date = monthly_sum_df['Year and Month'].iloc[i]
            month_rain = (monthly_sum_df['Monthly Rainfall'][monthly_sum_df['Year and Month'] == m_date])
            month_sum_rain = float(month_rain)
            august.append(month_sum_rain)

    monthly_df.insert(8, 'August', august)

    #12.4.9 Insert 'September' Column
    september = []
    for i in range(len(monthly_sum_df)):
        if monthly_sum_df['Year and Month'].iloc[i][5:]  == '09':
            m_date = monthly_sum_df['Year and Month'].iloc[i]
            month_rain = (monthly_sum_df['Monthly Rainfall'][monthly_sum_df['Year and Month'] == m_date])
            month_sum_rain = float(month_rain)
            september.append(month_sum_rain)

    monthly_df.insert(9, 'September', september)

    #12.4.10 Insert 'October' Column
    october = []
    for i in range(len(monthly_sum_df)):
        if monthly_sum_df['Year and Month'].iloc[i][5:]  == '10':
            m_date = monthly_sum_df['Year and Month'].iloc[i]
            month_rain = (monthly_sum_df['Monthly Rainfall'][monthly_sum_df['Year and Month'] == m_date])
            month_sum_rain = float(month_rain)
            october.append(month_sum_rain)

    monthly_df.insert(10, 'October', october)

    #12.4.11 Insert 'November' Column
    november = []
    for i in range(len(monthly_sum_df)):
        if monthly_sum_df['Year and Month'].iloc[i][5:]  == '11':
            m_date = monthly_sum_df['Year and Month'].iloc[i]
            month_rain = (monthly_sum_df['Monthly Rainfall'][monthly_sum_df['Year and Month'] == m_date])
            month_sum_rain = float(month_rain)
            november.append(month_sum_rain)

    monthly_df.insert(11, 'November', november)

    #12.4.12 Insert 'December' Column
    december = []
    for i in range(len(monthly_sum_df)):
        if monthly_sum_df['Year and Month'].iloc[i][5:]  == '12':
            m_date = monthly_sum_df['Year and Month'].iloc[i]
            month_rain = (monthly_sum_df['Monthly Rainfall'][monthly_sum_df['Year and Month'] == m_date])
            month_sum_rain = float(month_rain)
            december.append(month_sum_rain)

    monthly_df.insert(12, 'December', december)

    #**************************************************************************************************
    #12.5 Insert 'Total Rainfall (mm) in Year' Column
    total_rain_year = []
    for i in range(len(monthly_df)):
        yearly_sum_rain = (monthly_df.iloc[i].sum())-(monthly_df['Year'].iloc[i])
        total_rain_year.append(yearly_sum_rain)

    monthly_df.insert(13, 'Total Rainfall (mm) in Year', total_rain_year)

    ################################################################################################
    ############################ 13. Create 'monthlystats_df' Dataframe ############################
    ################################################################################################
    #13.1 Create 'monthlystats_df' Dataframe
    monthlystats_df = pd.DataFrame()

    #**************************************************************************************************
    #13.2 Insert 'Month' Column
    month_index = ['January', 'February', 'March', 'April', 'May', 'June', 'July', 'August', 'September', 'October', 'November', 'December']

    monthlystats_df.insert(0, 'Months', month_index)

    #**************************************************************************************************
    #13.3 Insert 'Average' Column

    monthly_len = float(len(monthly_df))

    jan1 = ((monthly_df['January'].sum())/monthly_len).round(decimals = 2)
    feb1 = ((monthly_df['February'].sum())/monthly_len).round(decimals = 2)
    mch1 = ((monthly_df['March'].sum())/monthly_len).round(decimals = 2)
    apr1 = ((monthly_df['April'].sum())/monthly_len).round(decimals = 2)
    may1 = ((monthly_df['May'].sum())/monthly_len).round(decimals = 2)
    jun1 = ((monthly_df['June'].sum())/monthly_len).round(decimals = 2)
    jul1 = ((monthly_df['July'].sum())/monthly_len).round(decimals = 2)
    aug1 = ((monthly_df['August'].sum())/monthly_len).round(decimals = 2)
    sep1 = ((monthly_df['September'].sum())/monthly_len).round(decimals = 2)
    octb1 = ((monthly_df['October'].sum())/monthly_len).round(decimals = 2)
    nov1 = ((monthly_df['November'].sum())/monthly_len).round(decimals = 2)
    dec1 = ((monthly_df['December'].sum())/monthly_len).round(decimals = 2)

    average = [jan1, feb1, mch1, apr1, may1, jun1, jul1, aug1, sep1, octb1, nov1, dec1]

    monthlystats_df.insert(1, 'Average', average)

    #**************************************************************************************************
    #13.4 Insert 'Standard Deviation (Sample Population)' Column
    jan2 = (statistics.stdev(monthly_df['January']))
    feb2 = (statistics.stdev(monthly_df['February']))
    mch2 = (statistics.stdev(monthly_df['March']))
    apr2 = (statistics.stdev(monthly_df['April']))
    may2 = (statistics.stdev(monthly_df['May']))
    jun2 = (statistics.stdev(monthly_df['June']))
    jul2 = (statistics.stdev(monthly_df['July']))
    aug2 = (statistics.stdev(monthly_df['August']))
    sep2 = (statistics.stdev(monthly_df['September']))
    octb2 = (statistics.stdev(monthly_df['October']))
    nov2 = (statistics.stdev(monthly_df['November']))
    dec2 = (statistics.stdev(monthly_df['December']))

    stdsp = [jan2, feb2, mch2, apr2, may2, jun2, jul2, aug2, sep2, octb2, nov2, dec2]

    monthlystats_df.insert(2, 'Standard Deviation (Sample Population)', stdsp)

    #**************************************************************************************************
    #13.5 Insert 'Standard Deviation (Entire Population)' Column
    jan3 = (statistics.pvariance(monthly_df['January']))
    feb3 = (statistics.pvariance(monthly_df['February']))
    mch3 = (statistics.pvariance(monthly_df['March']))
    apr3 = (statistics.pvariance(monthly_df['April']))
    may3 = (statistics.pvariance(monthly_df['May']))
    jun3 = (statistics.pvariance(monthly_df['June']))
    jul3 = (statistics.pvariance(monthly_df['July']))
    aug3 = (statistics.pvariance(monthly_df['August']))
    sep3 = (statistics.pvariance(monthly_df['September']))
    octb3 = (statistics.pvariance(monthly_df['October']))
    nov3 = (statistics.pvariance(monthly_df['November']))
    dec3 = (statistics.pvariance(monthly_df['December']))

    stdep = [jan3, feb3, mch3, apr3, may3, jun3, jul3, aug3, sep3, octb3, nov3, dec3]

    monthlystats_df.insert(3, 'Standard Deviation (Entire Population)', stdep)

    #**************************************************************************************************
    #13.6 Insert 'Mean' Column
    jan4 = (statistics.mean(monthly_df['January']))
    feb4 = (statistics.mean(monthly_df['February']))
    mch4 = (statistics.mean(monthly_df['March']))
    apr4 = (statistics.mean(monthly_df['April']))
    may4 = (statistics.mean(monthly_df['May']))
    jun4 = (statistics.mean(monthly_df['June']))
    jul4 = (statistics.mean(monthly_df['July']))
    aug4 = (statistics.mean(monthly_df['August']))
    sep4 = (statistics.mean(monthly_df['September']))
    octb4 = (statistics.mean(monthly_df['October']))
    nov4 = (statistics.mean(monthly_df['November']))
    dec4 = (statistics.mean(monthly_df['December']))

    mean = [jan4, feb4, mch4, apr4, may4, jun4, jul4, aug4, sep4, octb4, nov4, dec4]

    monthlystats_df.insert(3, 'Mean', mean)

    ################################################################################################
    ############################# 14. Create 'lpiii_table_df' Dataframe ############################
    ################################################################################################
    #14.1 Create 'pearson_wt_df' Dataframe
    #Read Pearson Type (III) WT Values Excel Spreadsheet
    pearson_wt_df = pd.read_excel(python_path + 'pearson_typeiii_wt_values.xlsx')

    #Change '1.0101' value to '1' in order to do calculations
    pearson_wt_df.loc[0].at['Return Period (Years)']=1

    #**************************************************************************************************
    #14.2 Create 'lpiii_table_df' Dataframe
    #14.2.1 Create 'lpiii_table_df' Dataframe
    lpiii_table_df = pd.DataFrame()

    #14.2.2 Insert 'Year' Column
    lpiii_table_df.insert(0, 'Year', gumbelest_df['Year'])

    #14.2.3 Insert 'Observed Max (mm/d)' Column
    lpiii_table_df.insert(1, 'Observed Max (mm/d)', gumbelest_df['Max 1-Day Rainfall (mm)'])

    #14.2.4 Insert 'i (rank)' Column
    lpiii_table_df.insert(2, 'i (Rank)', gumbelest_df['i (Rank)'])

    #14.2.5 Insert 'Log Observed Max (mm)' Column
    log_obsmax = []
    for i in range(len(lpiii_table_df)):
        log_obsmax.append(math.log10(lpiii_table_df['Observed Max (mm/d)'].iloc[i]))

    lpiii_table_df.insert(3, 'Log Observed Max (mm)', log_obsmax)

    #14.2.6 Insert '(LogQ - avg(Log Q))^2' Column
    log_minusavg_2 = []
    for i in range(len(lpiii_table_df)):
        math_log1 = lpiii_table_df['Log Observed Max (mm)'].iloc[i]
        avg_log1 = ((lpiii_table_df['Log Observed Max (mm)'].sum())/(len(lpiii_table_df))).round(decimals = 2)
        log_minusavg_2.append(pow((math_log1 - avg_log1), 2))

    lpiii_table_df.insert(4, '(LogQ - avg(Log Q))^2', log_minusavg_2)

    #14.2.7 Insert '(LogQ - avg(Log Q))^3' Column
    log_minusavg_3 = []
    for i in range(len(lpiii_table_df)):
        math_log2 = lpiii_table_df['Log Observed Max (mm)'].iloc[i]
        avg_log2 = ((lpiii_table_df['Log Observed Max (mm)'].sum())/(len(lpiii_table_df))).round(decimals = 2)
        log_minusavg_3.append(pow((math_log2 - avg_log2), 3))

    lpiii_table_df.insert(5, '(LogQ - avg(Log Q))^3', log_minusavg_3)

    #14.2.8 Insert 'Return Period (T)' Column
    n_plus1 = len(lpiii_table_df) + 1

    retperiod_t = []
    for i in range(len(lpiii_table_df)):
        value_t = (n_plus1 / lpiii_table_df['i (Rank)'].iloc[i])
        retperiod_t.append(value_t)

    lpiii_table_df.insert(6, 'Return Period (T)', retperiod_t)

    #14.2.9 Insert 'Exceedence Probability (P)' Column
    exc_p = []
    for i in range(len(lpiii_table_df)):
        exc_p.append(1 / lpiii_table_df['Return Period (T)'].iloc[i])

    lpiii_table_df.insert(7, 'Exceedence Probability (P)', exc_p)

    #14.2.10 Insert 'Non-Exceedence Probability (q)' Column
    nonexc_q = []
    for i in range(len(lpiii_table_df)):
        nonexc_q.append(1 - lpiii_table_df['Exceedence Probability (P)'].iloc[i])

    lpiii_table_df.insert(8, 'Non-Exceedence Probability (q)', nonexc_q)

    #**************************************************************************************************
    #14.3 Create 'wt_calculations' DataFrame
    #14.3.1 Create 'wt_calculations' DataFrame
    wt_calc_table_df = pd.DataFrame()

    #14.3.2 Create Variables
    n_obs = len(lpiii_table_df)
    avg_log_q = (lpiii_table_df['Log Observed Max (mm)'].sum())/len(log_obsmax)
    variance = (lpiii_table_df['(LogQ - avg(Log Q))^2'].sum())/(n_obs+1)
    std_dev_slog = np.sqrt(variance)
    skew_coeff_g = ((n_obs)/((n_obs-1)*(n_obs-2)))*((lpiii_table_df['(LogQ - avg(Log Q))^3'].sum())/(np.power(std_dev_slog, 3)))

    #14.3.3 Insert 'Return Period T (Year)' Column
    t_year = ['0.5', '1', '1.2', '1.5', '2', '5', '10', '20', '25', '35', '50', '100', '200', '500', '1000']
    return_p_year = ['1', '2', '5', '10', '25','50', '100', '200', '1000']
    wt_calc_table_df.insert(0, 'Return Period T (year)', t_year)

    #14.3.4 Create 'wt_calculations_df' Dataframe
    #14.3.4.1 Create 'wt_calculations_df' Dataframe
    wt_calculations_df = pd.DataFrame()

    #14.3.4.2 Insert 'Return Period T (year)' Column
    wt_calculations_df.insert(0, 'Return Period T (year)', return_p_year)

    #14.3.4.3 Define the Variables needed to find the Wt Value
    wt_per_g = [3, 2.9, 2.8, 2.7, 2.6, 2.5, 2.4, 2.3, 2.2, 2.1, 2, 1.9, 1.8, 1.7, 1.6, 1.5, 1.4, 1.3, 1.2, 1.1, 1, 0.9, 0.8, 0.7, 0.6, 0.5, 0.4, 0.3, 0.2, 0.1, 0, -0.1, -0.2, -0.3, -0.4, -0.5, -0.6, -0.7, -0.8, -0.9, -1, -1.1, -1.2, -1.3, -1.4, -1.5, -1.6, -1.7, -1.8, -1.9, -2, -2.1, -2.2, -2.3, -2.4, -2.5, -2.6, -2.7, -2.8, -2.9, -3]

    rp1 = None
    for val in wt_per_g:
        if val > skew_coeff_g:
            if rp1 is None or val < rp1:
                rp1 = val
            else:
                break

    rp2 = None
    for varsl in wt_per_g:
        if val < skew_coeff_g:
            if rp2 is None or val > rp2:
                rp2 = val
            else:
                break

    #14.3.4.3 Insert 'Return Period (Years) 1' Column
    #Insert a row where the value in the 'Return Period (Years)' row corresponds with the value in the column where the number is closest to the skew_coeff_g, but smaller.
    rp1_wt_value = []
    for i in range(len(pearson_wt_df['Return Period (Years)'])):
        wt_value_rp1 = pearson_wt_df[rp1].iloc[i]
        rp1_wt_value.append(wt_value_rp1)

    wt_calculations_df.insert(1, rp1, rp1_wt_value)

    #14.3.4.4 Insert 'Return Period (Years) 2' Column
    #Insert a row where the value in the 'Return Period (Years)' row corresponds with the value in the column where the number is closest to the skew_coeff_g, but larger.
    rp2_wt_value = []
    for i in range(len(pearson_wt_df['Return Period (Years)'])):
        wt_value_rp2 = pearson_wt_df[rp2].iloc[i]
        rp2_wt_value.append(wt_value_rp2)

    wt_calculations_df.insert(2, rp2, rp2_wt_value)

    #14.3.4.5 Insert 'Wt Value' Column
    final_wt_value = []
    for i in range(len(pearson_wt_df['Return Period (Years)'])):
        final_wt_value.append(rp2_wt_value[i] +(((rp1_wt_value[i]-rp2_wt_value[i])/(rp1-rp2))*(skew_coeff_g-rp2)))

    wt_calculations_df.insert(3, 'Wt Value', final_wt_value)

    #14.3.5 Insert 'Wt Value' Column
    #14.3.5.1 Define Variables
    u = wt_calculations_df['Wt Value']

    #return period t (year) = 0.5
    return_0_5_wt = (0.5-1.0101)*((u.iloc[0]-u.iloc[1])/(1.0101-2))+u.iloc[0]
    #return period t (year) = 1.5
    return_1_5_wt = u.iloc[0]+((u.iloc[1]-u.iloc[0])/(2-1.0101))*(1.5-1.0101)
    #return period t (year) = 1.2
    return_1_2_wt = u.iloc[0]+((return_1_5_wt-u.iloc[0])/(1.5-1.0101))*(1.2-1.0101)
    #return period t (year) = 20
    return_20_wt = u.iloc[3]+((u.iloc[4]-u.iloc[3])/(25-10))*(20-10)
    #return period t (year) = 35
    return_35_wt = u.iloc[4]+((u.iloc[5]-u.iloc[4])/(50-25))*(35-25)
    #return period t (year) = 500
    return_500_wt = u.iloc[7]+((u.iloc[8]-u.iloc[7])/(1000-200))*(500-200)

    wt_return_period = [return_0_5_wt, u.iloc[0], return_1_2_wt, return_1_5_wt, u.iloc[1],
                        u.iloc[2], u.iloc[3], return_20_wt, u.iloc[4], return_35_wt,
                        u.iloc[5], u.iloc[6], u.iloc[7], return_500_wt, u.iloc[8]]

    #14.3.5.2 Insert 'Wt Value' Column
    wt_calc_table_df.insert(1, 'Wt Value', wt_return_period)

    #14.3.6 Insert 'Log Qt' Column
    log_qt_wt = []
    for i in range(len(wt_calc_table_df)):
        log_value = avg_log_q+(std_dev_slog*wt_calc_table_df['Wt Value'].iloc[i])
        log_qt_wt.append(log_value)

    wt_calc_table_df.insert(2, 'Log Qt', log_qt_wt)

    #14.3.7 Insert 'Pearson Estimate (mm/day)' Column
    pearson_est_wt = []
    for i in range(len(wt_calc_table_df)):
        pearson_est_wt.append(np.power(10, wt_calc_table_df['Log Qt'].iloc[i]))

    wt_calc_table_df.insert(3, 'Pearson Estimate (mm/day)', pearson_est_wt)

    #14.3.8 Insert 'Full Series (mm/day)' Column
    wt_calc_table_df['Return Period T (year)'] = wt_calc_table_df['Return Period T (year)'].astype(float)

    fullgumbel_df['Full Series (mm/d)'] = fullgumbel_df['Full Series (mm/d)'].replace('', 0.00)

    full_series_est_wt = []
    for i in range(len(wt_calc_table_df)):
        if wt_calc_table_df['Return Period T (year)'].iloc[i] in fullgumbel_df['Return Period T (Year)'].values:
            ret_period_value = wt_calc_table_df['Return Period T (year)'].iloc[i]
            full_series_value = (fullgumbel_df['Full Series (mm/d)'][fullgumbel_df['Return Period T (Year)'] == ret_period_value])
            full_series_value_wt = float(full_series_value)
            full_series_est_wt.append(full_series_value_wt)
        else:
            full_series_est_wt.append(np.nan)

    wt_calc_table_df.insert(4, 'Full Series (mm/day)', full_series_est_wt)

    #14.2.11 Insert 'Wt' Column
    lpiii_table_df['Return Period (T)'] = lpiii_table_df['Return Period (T)'].astype(float)
    wt_calc_table_df['Return Period T (year)'] = wt_calc_table_df['Return Period T (year)'].astype(float)

    standard_variate_wt = []
    for i in range(len(lpiii_table_df)):
        if lpiii_table_df['Return Period (T)'].iloc[i] <= wt_calc_table_df['Return Period T (year)'].max():
            return_period = lpiii_table_df['Return Period (T)'].iloc[i]
            return_period_ref = wt_calc_table_df['Return Period T (year)']
            wt_value_ref = wt_calc_table_df['Wt Value']
            for j in range(len(wt_calc_table_df)):
                if (return_period_ref.iloc[j] <= return_period) and (return_period_ref.iloc[j+1] >= return_period):
                    #Referenced Return Period (T) Values
                    rpt2 = return_period_ref.iloc[j]
                    rpt1 = return_period_ref.iloc[j+1]
                    #Corresponding Wt Values
                    wt_rpt2 = wt_value_ref.iloc[j]
                    wt_rpt1 = wt_value_ref.iloc[j+1]
                    break
            wt_formula = wt_rpt2 + (((wt_rpt1-wt_rpt2) / (rpt1-rpt2)) * (return_period-rpt2))
            standard_variate_wt.append(wt_formula)
        else:
            standard_variate_wt.append('n/a')

    lpiii_table_df.insert(9, 'Wt', standard_variate_wt)

    #14.2.13 Insert 'LP (III) Estimate (mm/d)' Column
    lp_iii_est_mmd = []
    for i in range(len(lpiii_table_df)):
        lp_est_calc = np.power(10, (avg_log_q+(std_dev_slog * (lpiii_table_df['Wt'].iloc[i]))))
        lp_iii_est_mmd.append(lp_est_calc)

    lpiii_table_df.insert(10, 'LP (III) Estimate (mm/d)', lp_iii_est_mmd)

    ################################################################################################
    ############################### 15. Create 'gevmev_df' Dataframe ###############################
    ################################################################################################
    #15.1 Create 'gevmev_df' Dataframe
    gevmev_df = pd.DataFrame()

    #**************************************************************************************************
    #15.2 Insert 'Return Period (T)' Column
    gevmev_df.insert(0, 'Return Period (T)', fullgumbel_df['Return Period T (Year)'])

    #**************************************************************************************************
    #15.3 Insert 'Full Series (mm/d)' Column
    gevmev_df.insert(1, 'Full Series (mm/d)', fullgumbel_df['Full Series (mm/d)'])
    gevmev_df['Full Series (mm/d)'].replace([0], '', inplace=True)

    #**************************************************************************************************
    #15.4 Insert 'Gumbel Estimate (mm/d)' Column
    gevmev_df.insert(2, 'Gumbel Estimate (mm/d)', fullgumbel_df['Gumbel Estimate (mm/d)'])

    #**************************************************************************************************
    #15.5 Insert 'Log Pearson (III) Estimate (mm/d)' Column
    gevmev_df.insert(3, 'Log Pearson (III) Estimate (mm/d)', wt_calc_table_df['Pearson Estimate (mm/day)'])

    #**************************************************************************************************
    #15.6 Insert 'GEV Estimate (mm/d)' Column
    #15.6.1 Create 'gev_estimte_calculation_df' DataFrame
    gev_estimte_calculation_df = pd.DataFrame()

    #15.6.2 Insert 'time' Column
    gev_estimte_calculation_df.insert(0,'time', og_rainfall['Date'])

    #15.6.3 Insert 'PRCP' Column
    gev_estimte_calculation_df.insert(1,'PRCP', og_rainfall['Rainfall (mm/d)'])

    #15.6.4 Insert 'YEAR' Column
    year_list = []
    for i in range(len(og_rainfall)):
        year_number = int(og_rainfall['Date'].iloc[i][:4])
        year_list.append(int(year_number))

    gev_estimte_calculation_df.insert(2,'YEAR', year_list)

    #15.6.5 Change Data Types
    gev_estimte_calculation_df['PRCP'] = gev_estimte_calculation_df['PRCP'].astype(float)

    #15.6.6 Change 'time' Column to index
    gev_estimte_calculation_df = gev_estimte_calculation_df.set_index('time')

    #15.6.7 Define Variables Returns Periods of Interest
    #Return periods of interest
    #The next step is to specify the return periods we are interested in. Create two numpy arrays. The first array contains the return periods ($T$) we are interested in. These are the same as you used in week two: 0.5, 1, 1.2, 1.5, 2, 5, 10, 20, 35, 50, 100, 200, 500 and 1000 years.
    #The second array has the annual probability of non-exceedance of certain precipitation event ($F$).
    T = np.asarray([0.5, 1, 1.2, 1.5, 2, 5, 10, 20, 35, 50, 100, 200, 500, 1000])
    F = np.exp(-(1/T))*(1/T)**0

    #15.6.8 Calculate AMS, 1, T_a, T_observations
    #Just as for the Gumbel exercise last week, in order to compute the GEV we need the annual maxima series (AMS) for our precipitation data. The package mevpy has a function for this: __mev.tab_rain_max(df)__.
    #This function not only computes the AMS, but also the plotting position estimate of their non-exceedance probability ($q$) and their relative return times ($T_a$).
    #We also calculate the observed return periods (__T_observations__), and we make a dataframe with all these variables for visualisation.
    AMS, q, T_a    = mev.tab_rain_max(gev_estimte_calculation_df)
    T_observations = -1/(np.log(1-(1/T_a)))

    df_ams         = pd.DataFrame(data=[AMS, q, T_a, T_observations]).T
    df_ams.columns = ['AMS', 'q', 'T_a', 'T_observations']

    df_ams.insert(0,'Year', monthly_df['Year'])

    #15.6.9 Calculate GEV Parameters: Shape (csi ($\xi$)), Scale (sigma ($\sigma$)) & Location (mu  ($\mu$))

    #The GEV has three parameters:
    # - GEV shape parameter    = csi ($\xi$)
    # - GEV scale parameter    = sigma ($\sigma$)
    # - GEV location parameter = mu  ($\mu$)
    #The function __mev.gev_fit(AMS)__ estimates the GEV parameters automatically based on L-moments.
    csi, sigma, mu = mev.gev_fit(AMS)

    #15.6.10 Use the function __mev.gev_quant($F$, $\xi$, $\sigma$, $\mu$)__ to compute the GEV estimate, and create a dataframe to easily visualise this.
    #Now we are interested in the amount of precipitation corresponding to the different return periods of interest ($T$).
    gev_estimate    = mev.gev_quant(F, csi, sigma, mu)

    df_gev_estimate = pd.DataFrame(gev_estimate,columns=['GEV estimate [mm/day]'])
    df_gev_estimate.index = T # Return Period

    #15.6.11 Insert 'gevmev_gevestimate' Column
    gevmev_gevestimate = []
    for i in range(len(df_gev_estimate)):
        gevmev_gevestimate.append(df_gev_estimate['GEV estimate [mm/day]'].iloc[i])

    gevmev_df.insert(4, 'GEV Estimate (mm/d)', gevmev_gevestimate)

    #**************************************************************************************************
    #15.7 Insert 'MEV Estimate (mm/d)' Column
    #The MEV is the Metastatistical Extreme Value distribution. The MEV uses the full distribution of the data, so it also includes the information of the "ordinary" rainfall events to estimate the extremes. The Weibull distribution is fitted to the annual data, whereafter the MEV is computed.
    #15.7.1 Calculate the Parameters N (Dimension of the sample), C (Weibull scale parameter), W (Weibull shape parameter)

    #The MEV has the following parameters:
    # - Dimension of the sample = $N$ ( = number of wet days, i.e. above the threshold of 1 mm/day)
    # - Weibull scale parameter = $C$
    # - Weibull shape parameter = $W$
    #The function __mev.mev_fit(df, threshold=1)__ estimates the MEV parameters automatically based on Probability Weighthed Moments.
    N, C, W =  mev.mev_fit(gev_estimte_calculation_df, threshold=1)

    #15.7.2 Calculate MEV Estimates
    #Compute the amounts of precipitation corresponding to the different return periods of interest($T$).
    #Use the function __mev.mev_quant($q$, $x0$, $N$, $C$, $W$)__ to calculate the MEV estimate.
    #Note that the MEV does not have an analytical solution, therefore, we must compute it numerically.
    #The parameter $x0$ is the starting guess for the numerical solution of the MEV.
    #50 mm/day is a good initial guess for $x0$.
    sorted_df = max1day_df.sort_values(by='Max 1-Day Rainfall (mm)', ascending=True)
    x0_initial = sorted_df['Max 1-Day Rainfall (mm)'].iloc[0]
    x0 = np.floor(x0_initial)

    gevmev_mevestimate =  mev.mev_quant(F, x0, N, C, W)

    #15.7.3 Insert 'MEV Estimate (mm/d)' Column
    gevmev_df.insert(5, 'MEV Estimate (mm/d)', gevmev_mevestimate)

    #**************************************************************************************************
    #15.8 Create 'mev_values_df' and  'gev_values_df' Dataframe
    #15.8.1 Insert 'csi', 'sigma' and 'mu' Column
    gev_values_df         = pd.DataFrame(data=[csi, sigma, mu]).T
    gev_values_df.columns = ['csi', 'sigma', 'mu']

    #15.8.2 Insert 'N', 'C' and 'W' Columns
    mev_values_df         = pd.DataFrame(data=[N, C, W]).T
    mev_values_df.columns = ['N', 'C', 'W' ]

    #15.8.3 Insert 'Year' Columns
    mev_values_df.insert(0,'Year', monthly_df['Year'])

    #15.8.5 Display 'gev_values_df' Dataframe
    gev_values_df['x0 (starting guess of MEV)']=x0

    ################################################################################################
    ########################## 16. Create 'stochastic_table_df' Dataframe ##########################
    ################################################################################################
    #16.1 Sort Gumble Estimate Dataframe by Year Values
    gumbelest_stoch = gumbelest_df.sort_values(['Year'])

    #**************************************************************************************************
    #16.2 Create 'stochastic_table_df' Dataframe
    #16.2.1 Create 'stochastic_table_df' Dataframe
    stochastic_table_df = pd.DataFrame()

    #16.2.2 Insert 'Year' Column
    stochastic_table_df.insert(0, 'Year', gumbelest_stoch['Year'])

    #16.2.3 Insert 'Annual Rainfall (mm)' Column
    rain_year = []
    for i in range(len(gumbelest_stoch)):
        sum_rain_gumbelest = year_df.groupby('Year')['Rainfall (mm/d)'].sum()
        rain_year.append(sum_rain_gumbelest.iloc[i])

    stochastic_table_df.insert(1, 'Annual Rainfall (mm)', rain_year)

    #16.2.4 Insert 'p (Exceedance)' Column
    stochastic_table_df.insert(2, 'p (Exceedance)', gumbelest_stoch['P (Exceedance)'])

    #16.2.5 Insert 'q (Non-Exceedance)' Column
    stochastic_table_df.insert(3, 'q (Non-Exceedance)', gumbelest_stoch['Q (Non-Exc)'])

    #16.2.6 Insert 'Rank' Column
    stochastic_table_df = stochastic_table_df.sort_values(['Annual Rainfall (mm)'],
                  ascending = [False])

    rank_stochastic = []
    for i in range(len(stochastic_table_df)):
        rank_stochastic.append(i+1)

    stochastic_table_df.insert(2, 'Rank', rank_stochastic)

    #16.2.7 Insert 'T (Return Period)' Column
    stochastic_table_df = stochastic_table_df.sort_values(['Year'],
                  ascending = [True])

    stochastic_table_df.insert(3, 'T (Return Period)', gumbelest_df['T_a (Return Period)'])

    #16.2.8 Insert 'y' Column
    stochastic_table_df.insert(4, 'Y (Return Var)', gumbelest_df['Y (Return Var)'])

    #16.2.9 Insert 'Gumbel Estimate (mm/d)' Column
    stochastic_table_df.insert(5, 'Gumbel Estimate (mm/d)', gumbelest_df['Gumbel Estimate (mm/d)'])

    #16.2.10 Insert 'Log Observed Max (mm)' Column
    stochastic_table_df.insert(6, 'Log Observed Max (mm)', lpiii_table_df['Log Observed Max (mm)'])

    #16.2.11 Insert '(LogQ - avg(LogQ))^2' Column
    stochastic_table_df.insert(7, '(LogQ - avg(Log Q))^2', lpiii_table_df['(LogQ - avg(Log Q))^2'])

    #16.2.12 Insert '(LogQ - avg(LogQ))^3' Column
    stochastic_table_df.insert(8, '(LogQ - avg(Log Q))^3', lpiii_table_df['(LogQ - avg(Log Q))^3'])

    #16.2.13 Insert 'Exceedence Probability (P)' Column
    stochastic_table_df.insert(9, 'Exceedence Probability (P)', lpiii_table_df['Exceedence Probability (P)'])

    #16.2.14 Insert 'Non-Exceedence Probability (q)' Column
    stochastic_table_df.insert(10, 'Non-Exceedence Probability (q)', lpiii_table_df['Non-Exceedence Probability (q)'])

    #16.2.15 Insert 'Wt' Column
    stochastic_table_df.insert(11, 'Wt', lpiii_table_df['Wt'])


    #**************************************************************************************************
    #**************************************************************************************************
    #************************************ CREATE EXCEL SPREADSHEET ************************************
    #**************************************************************************************************
    #**************************************************************************************************
    #1. Save all the DataFrames to an Excel Spreadsheet
    empty_df = pd.DataFrame()

    dataframe0 = empty_df
    dataframe1 = df
    dataframe2 = oneday_df
    dataframe3 = ddf_df
    dataframe4 = gumbelest_df
    dataframe5 = fullgumbel_df
    dataframe6 = resmass_df
    dataframe7 = monthly_df
    dataframe8 = lpiii_table_df
    dataframe9 = gevmev_df

    with pd.ExcelWriter('output_file.xlsx', engine='xlsxwriter') as writer:
        dataframe0.to_excel(writer, sheet_name='Introduction', index=False)
        dataframe1.to_excel(writer, sheet_name='Original Rainfall Data', index=False)
        dataframe2.to_excel(writer, sheet_name='One Day Analyses', index=False)
        dataframe3.to_excel(writer, sheet_name='One Day Probability Data Frame', index=False)
        dataframe4.to_excel(writer, sheet_name='Gumbel Estimate (mm_day)', index=False)
        dataframe5.to_excel(writer, sheet_name='Gumbel Est and Full Series', index=False)
        dataframe6.to_excel(writer, sheet_name='Residual Mass', index=False)
        dataframe7.to_excel(writer, sheet_name='Monthly Rainfall (Box_Whisker)', index=False)
        dataframe0.to_excel(writer, sheet_name='Log Pearson (III)', index=False)
        dataframe0.to_excel(writer, sheet_name='GEV and MEV', index=False)
        dataframe8.to_excel(writer, sheet_name='LP (III) Estimate', index=False)
        dataframe9.to_excel(writer, sheet_name='LP Est, GEV and MEV', index=False)

    print('You have successfully saved your DataFrames to an Excel Spreadsheet')

    #**************************************************************************************************
    #2. Export Tables
    #Wt Value Table
    table8_1 = wt_calculations_df
    dfi.export(table8_1,"wt_value_table.png")

    #Pearson Est Table
    table8_2 = wt_calc_table_df
    dfi.export(table8_2,"pearson_est_table.png")

    #GEV Variables_01
    if len(df_ams) < 100:
        table9_1 = df_ams
        dfi.export(table9_1,"mev_var_01.png")
    else:
        table9_1 = df_ams.head(50)
        dfi.export(table9_1,"mev_var_01.png")

    #MEV Variables_02
    table9_2 = gev_values_df
    dfi.export(table9_2,"mev_var_02.png")

    #MEV Variables_03
    if len(df_ams) < 100:
        table9_3 = mev_values_df
        dfi.export(table9_3,"mev_var_03.png")
    else:
        table9_3 = mev_values_df.head(50)
        dfi.export(table9_3,"mev_var_03.png")

    time.sleep(5)

    #**************************************************************************************************
    #3. Modify Excel Spreadsheet
    #3.1 Change Data Types
    wb = load_workbook('output_file.xlsx')
    ws = wb['Original Rainfall Data']

    date_change = []
    for i in range(len(df)+2):
        if i != 0:
            date_change.append('A' + str(i))

    for i in range(len(date_change)):
            ws[(date_change[i])].number_format = 'YYYY-MM-DD'

    wb.save('output_file.xlsx')
    wb.close()

    #3.2 Change Row width to fit Contents
    wb = load_workbook('output_file.xlsx')
    ws = wb['Original Rainfall Data']
    ws.column_dimensions['A'].width = 11.5
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 25
    ws.column_dimensions['E'].width = 25
    ws.column_dimensions['F'].width = 25
    wb.save('output_file.xlsx')
    wb.close()

    wb = load_workbook('output_file.xlsx')
    ws = wb['One Day Analyses']
    ws.column_dimensions['A'].width = 24
    ws.column_dimensions['B'].width = 16
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['D'].width = 15.5
    ws.column_dimensions['E'].width = 13
    ws.column_dimensions['F'].width = 13
    wb.save('output_file.xlsx')
    wb.close()

    wb = load_workbook('output_file.xlsx')
    ws = wb['One Day Probability Data Frame']
    ws.column_dimensions['A'].width = 27.5
    ws.column_dimensions['B'].width = 22
    wb.save('output_file.xlsx')
    wb.close()

    wb = load_workbook('output_file.xlsx')
    ws = wb['Gumbel Estimate (mm_day)']
    ws.column_dimensions['A'].width = 6.5
    ws.column_dimensions['B'].width = 29.5
    ws.column_dimensions['C'].width = 11.5
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 19.5
    ws.column_dimensions['F'].width = 15.5
    ws.column_dimensions['G'].width = 24.5
    ws.column_dimensions['H'].width = 17.5
    ws.column_dimensions['I'].width = 30
    wb.save('output_file.xlsx')
    wb.close()

    wb = load_workbook('output_file.xlsx')
    ws = wb['Gumbel Est and Full Series']
    ws.column_dimensions['A'].width = 27.5
    ws.column_dimensions['B'].width = 16.5
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['D'].width = 30
    ws.column_dimensions['E'].width = 22.5
    wb.save('output_file.xlsx')
    wb.close()

    wb = load_workbook('output_file.xlsx')
    ws = wb['Residual Mass']
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 30
    ws.column_dimensions['D'].width = 30
    ws.column_dimensions['E'].width = 50
    ws.column_dimensions['F'].width = 60
    wb.save('output_file.xlsx')
    wb.close()

    wb = load_workbook('output_file.xlsx')
    ws = wb['Monthly Rainfall (Box_Whisker)']
    ws.column_dimensions['A'].width = 6.5
    ws.column_dimensions['B'].width = 13.5
    ws.column_dimensions['C'].width = 13.5
    ws.column_dimensions['D'].width = 13.5
    ws.column_dimensions['E'].width = 13.5
    ws.column_dimensions['F'].width = 13.5
    ws.column_dimensions['G'].width = 13.5
    ws.column_dimensions['H'].width = 13.5
    ws.column_dimensions['I'].width = 13.5
    ws.column_dimensions['J'].width = 13.5
    ws.column_dimensions['K'].width = 13.5
    ws.column_dimensions['L'].width = 13.5
    ws.column_dimensions['M'].width = 13.5
    ws.column_dimensions['N'].width = 33
    wb.save('output_file.xlsx')
    wb.close()

    wb = load_workbook('output_file.xlsx')
    ws = wb['LP (III) Estimate']
    ws.column_dimensions['A'].width = 6.5
    ws.column_dimensions['B'].width = 28
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 28
    ws.column_dimensions['E'].width = 28
    ws.column_dimensions['F'].width = 28
    ws.column_dimensions['G'].width = 21.5
    ws.column_dimensions['H'].width = 38
    ws.column_dimensions['I'].width = 38
    ws.column_dimensions['J'].width = 14
    ws.column_dimensions['K'].width = 29
    wb.save('output_file.xlsx')
    wb.close()

    wb = load_workbook('output_file.xlsx')
    ws = wb['LP Est, GEV and MEV']
    ws.column_dimensions['A'].width = 22.5
    ws.column_dimensions['B'].width = 22.5
    ws.column_dimensions['C'].width = 30
    ws.column_dimensions['D'].width = 41
    ws.column_dimensions['E'].width = 27
    ws.column_dimensions['F'].width = 27
    wb.save('output_file.xlsx')
    wb.close()

    #3.3 Change Top Row Colour to Dark Blue
    wb = load_workbook('output_file.xlsx')
    ws = wb['Original Rainfall Data']
    ws['A1'].fill = PatternFill(start_color = '5F779D', end_color = '5F779D', fill_type = 'solid')
    ws['B1'].fill = PatternFill(start_color = '5F779D', end_color = '5F779D', fill_type = 'solid')
    ws['C1'].fill = PatternFill(start_color = '5F779D', end_color = '5F779D', fill_type = 'solid')
    ws['D1'].fill = PatternFill(start_color = '5F779D', end_color = '5F779D', fill_type = 'solid')
    ws['E1'].fill = PatternFill(start_color = '5F779D', end_color = '5F779D', fill_type = 'solid')
    ws['F1'].fill = PatternFill(start_color = '5F779D', end_color = '5F779D', fill_type = 'solid')
    wb.save('output_file.xlsx')
    wb.close()

    wb = load_workbook('output_file.xlsx')
    ws = wb['One Day Analyses']
    ws['A1'].fill = PatternFill(start_color = '5F779D', end_color = '5F779D', fill_type = 'solid')
    ws['B1'].fill = PatternFill(start_color = '5F779D', end_color = '5F779D', fill_type = 'solid')
    ws['C1'].fill = PatternFill(start_color = '5F779D', end_color = '5F779D', fill_type = 'solid')
    ws['D1'].fill = PatternFill(start_color = '5F779D', end_color = '5F779D', fill_type = 'solid')
    ws['E1'].fill = PatternFill(start_color = '5F779D', end_color = '5F779D', fill_type = 'solid')
    ws['F1'].fill = PatternFill(start_color = '5F779D', end_color = '5F779D', fill_type = 'solid')
    wb.save('output_file.xlsx')
    wb.close()

    wb = load_workbook('output_file.xlsx')
    ws = wb['One Day Probability Data Frame']
    ws['A1'].fill = PatternFill(start_color = '5F779D', end_color = '5F779D', fill_type = 'solid')
    ws['B1'].fill = PatternFill(start_color = '5F779D', end_color = '5F779D', fill_type = 'solid')
    wb.save('output_file.xlsx')
    wb.close()

    wb = load_workbook('output_file.xlsx')
    ws = wb['Gumbel Estimate (mm_day)']
    ws['A1'].fill = PatternFill(start_color = '5F779D', end_color = '5F779D', fill_type = 'solid')
    ws['B1'].fill = PatternFill(start_color = '5F779D', end_color = '5F779D', fill_type = 'solid')
    ws['C1'].fill = PatternFill(start_color = '5F779D', end_color = '5F779D', fill_type = 'solid')
    ws['D1'].fill = PatternFill(start_color = '5F779D', end_color = '5F779D', fill_type = 'solid')
    ws['E1'].fill = PatternFill(start_color = '5F779D', end_color = '5F779D', fill_type = 'solid')
    ws['F1'].fill = PatternFill(start_color = '5F779D', end_color = '5F779D', fill_type = 'solid')
    ws['G1'].fill = PatternFill(start_color = '5F779D', end_color = '5F779D', fill_type = 'solid')
    ws['H1'].fill = PatternFill(start_color = '5F779D', end_color = '5F779D', fill_type = 'solid')
    ws['I1'].fill = PatternFill(start_color = '5F779D', end_color = '5F779D', fill_type = 'solid')
    wb.save('output_file.xlsx')
    wb.close()

    wb = load_workbook('output_file.xlsx')
    ws = wb['Gumbel Est and Full Series']
    ws['A1'].fill = PatternFill(start_color = '5F779D', end_color = '5F779D', fill_type = 'solid')
    ws['B1'].fill = PatternFill(start_color = '5F779D', end_color = '5F779D', fill_type = 'solid')
    ws['C1'].fill = PatternFill(start_color = '5F779D', end_color = '5F779D', fill_type = 'solid')
    ws['D1'].fill = PatternFill(start_color = '5F779D', end_color = '5F779D', fill_type = 'solid')
    ws['E1'].fill = PatternFill(start_color = '5F779D', end_color = '5F779D', fill_type = 'solid')
    wb.save('output_file.xlsx')
    wb.close()

    wb = load_workbook('output_file.xlsx')
    ws = wb['Residual Mass']
    ws['A1'].fill = PatternFill(start_color = '5F779D', end_color = '5F779D', fill_type = 'solid')
    ws['B1'].fill = PatternFill(start_color = '5F779D', end_color = '5F779D', fill_type = 'solid')
    ws['C1'].fill = PatternFill(start_color = '5F779D', end_color = '5F779D', fill_type = 'solid')
    ws['D1'].fill = PatternFill(start_color = '5F779D', end_color = '5F779D', fill_type = 'solid')
    ws['E1'].fill = PatternFill(start_color = '5F779D', end_color = '5F779D', fill_type = 'solid')
    ws['F1'].fill = PatternFill(start_color = '5F779D', end_color = '5F779D', fill_type = 'solid')
    wb.save('output_file.xlsx')
    wb.close()

    wb = load_workbook('output_file.xlsx')
    ws = wb['Monthly Rainfall (Box_Whisker)']
    ws['A1'].fill = PatternFill(start_color = '5F779D', end_color = '5F779D', fill_type = 'solid')
    ws['B1'].fill = PatternFill(start_color = '5F779D', end_color = '5F779D', fill_type = 'solid')
    ws['C1'].fill = PatternFill(start_color = '5F779D', end_color = '5F779D', fill_type = 'solid')
    ws['D1'].fill = PatternFill(start_color = '5F779D', end_color = '5F779D', fill_type = 'solid')
    ws['E1'].fill = PatternFill(start_color = '5F779D', end_color = '5F779D', fill_type = 'solid')
    ws['F1'].fill = PatternFill(start_color = '5F779D', end_color = '5F779D', fill_type = 'solid')
    ws['G1'].fill = PatternFill(start_color = '5F779D', end_color = '5F779D', fill_type = 'solid')
    ws['H1'].fill = PatternFill(start_color = '5F779D', end_color = '5F779D', fill_type = 'solid')
    ws['I1'].fill = PatternFill(start_color = '5F779D', end_color = '5F779D', fill_type = 'solid')
    ws['J1'].fill = PatternFill(start_color = '5F779D', end_color = '5F779D', fill_type = 'solid')
    ws['K1'].fill = PatternFill(start_color = '5F779D', end_color = '5F779D', fill_type = 'solid')
    ws['L1'].fill = PatternFill(start_color = '5F779D', end_color = '5F779D', fill_type = 'solid')
    ws['M1'].fill = PatternFill(start_color = '5F779D', end_color = '5F779D', fill_type = 'solid')
    ws['N1'].fill = PatternFill(start_color = '5F779D', end_color = '5F779D', fill_type = 'solid')
    wb.save('output_file.xlsx')
    wb.close()

    wb = load_workbook('output_file.xlsx')
    ws = wb['LP (III) Estimate']
    ws['A1'].fill = PatternFill(start_color = '5F779D', end_color = '5F779D', fill_type = 'solid')
    ws['B1'].fill = PatternFill(start_color = '5F779D', end_color = '5F779D', fill_type = 'solid')
    ws['C1'].fill = PatternFill(start_color = '5F779D', end_color = '5F779D', fill_type = 'solid')
    ws['D1'].fill = PatternFill(start_color = '5F779D', end_color = '5F779D', fill_type = 'solid')
    ws['E1'].fill = PatternFill(start_color = '5F779D', end_color = '5F779D', fill_type = 'solid')
    ws['F1'].fill = PatternFill(start_color = '5F779D', end_color = '5F779D', fill_type = 'solid')
    ws['G1'].fill = PatternFill(start_color = '5F779D', end_color = '5F779D', fill_type = 'solid')
    ws['H1'].fill = PatternFill(start_color = '5F779D', end_color = '5F779D', fill_type = 'solid')
    ws['I1'].fill = PatternFill(start_color = '5F779D', end_color = '5F779D', fill_type = 'solid')
    ws['J1'].fill = PatternFill(start_color = '5F779D', end_color = '5F779D', fill_type = 'solid')
    ws['K1'].fill = PatternFill(start_color = '5F779D', end_color = '5F779D', fill_type = 'solid')
    wb.save('output_file.xlsx')
    wb.close()

    wb = load_workbook('output_file.xlsx')
    ws = wb['LP Est, GEV and MEV']
    ws['A1'].fill = PatternFill(start_color = '5F779D', end_color = '5F779D', fill_type = 'solid')
    ws['B1'].fill = PatternFill(start_color = '5F779D', end_color = '5F779D', fill_type = 'solid')
    ws['C1'].fill = PatternFill(start_color = '5F779D', end_color = '5F779D', fill_type = 'solid')
    ws['D1'].fill = PatternFill(start_color = '5F779D', end_color = '5F779D', fill_type = 'solid')
    ws['E1'].fill = PatternFill(start_color = '5F779D', end_color = '5F779D', fill_type = 'solid')
    ws['F1'].fill = PatternFill(start_color = '5F779D', end_color = '5F779D', fill_type = 'solid')
    wb.save('output_file.xlsx')
    wb.close()

    #3.4 Change every second Row Colour to Light Red/Light Maroon
    #3.4.1 Original Rainfall Data
    colourA = []
    for i in range(len(df)+2):
        if i % 2 == 0:
            if i != 0:
                colourA.append('A' + str(i))

    wb = load_workbook('output_file.xlsx')
    ws = wb['Original Rainfall Data']
    for i in range(len(colourA)):
        ws[(colourA[i])].fill = PatternFill(start_color = 'B6BED1', end_color = 'B6BED1', fill_type = 'solid')
    wb.save('output_file.xlsx')
    wb.close()

    colourB = []
    for i in range(len(df)+2):
        if i % 2 == 0:
            if i != 0:
                colourB.append('B' + str(i))

    wb = load_workbook('output_file.xlsx')
    ws = wb['Original Rainfall Data']
    for i in range(len(colourB)):
        ws[(colourB[i])].fill = PatternFill(start_color = 'B6BED1', end_color = 'B6BED1', fill_type = 'solid')

    wb.save('output_file.xlsx')
    wb.close()

    colourC = []
    for i in range(len(df)+2):
        if i % 2 == 0:
            if i != 0:
                colourC.append('C' + str(i))

    wb = load_workbook('output_file.xlsx')
    ws = wb['Original Rainfall Data']
    for i in range(len(colourC)):
        ws[(colourC[i])].fill = PatternFill(start_color = 'B6BED1', end_color = 'B6BED1', fill_type = 'solid')

    wb.save('output_file.xlsx')
    wb.close()

    colourD_extra = []
    for i in range(len(df)+2):
        if i % 2 == 0:
            if i != 0:
                colourD_extra.append('D' + str(i))

    wb = load_workbook('output_file.xlsx')
    ws = wb['Original Rainfall Data']
    for i in range(len(colourD_extra)):
        ws[(colourD_extra[i])].fill = PatternFill(start_color = 'B6BED1', end_color = 'B6BED1', fill_type = 'solid')

    wb.save('output_file.xlsx')
    wb.close()

    colourE_extra = []
    for i in range(len(df)+2):
        if i % 2 == 0:
            if i != 0:
                colourE_extra.append('E' + str(i))

    wb = load_workbook('output_file.xlsx')
    ws = wb['Original Rainfall Data']
    for i in range(len(colourE_extra)):
        ws[(colourE_extra[i])].fill = PatternFill(start_color = 'B6BED1', end_color = 'B6BED1', fill_type = 'solid')

    wb.save('output_file.xlsx')
    wb.close()

    colourF_extra = []
    for i in range(len(df)+2):
        if i % 2 == 0:
            if i != 0:
                colourF_extra.append('F' + str(i))

    wb = load_workbook('output_file.xlsx')
    ws = wb['Original Rainfall Data']
    for i in range(len(colourF_extra)):
        ws[(colourF_extra[i])].fill = PatternFill(start_color = 'B6BED1', end_color = 'B6BED1', fill_type = 'solid')

    wb.save('output_file.xlsx')
    wb.close()

    #3.4.2 One Day Analyses
    colourD = []
    for i in range(len(oneday_df)+2):
        if i % 2 == 0:
            if i != 0:
                colourD.append('A' + str(i))

    wb = load_workbook('output_file.xlsx')
    ws = wb['One Day Analyses']
    for i in range(len(colourD)):
        ws[(colourD[i])].fill = PatternFill(start_color = 'B6BED1', end_color = 'B6BED1', fill_type = 'solid')
    wb.save('output_file.xlsx')
    wb.close()

    colourE = []
    for i in range(len(oneday_df)+2):
        if i % 2 == 0:
            if i != 0:
                colourE.append('B' + str(i))

    wb = load_workbook('output_file.xlsx')
    ws = wb['One Day Analyses']
    for i in range(len(colourE)):
        ws[(colourE[i])].fill = PatternFill(start_color = 'B6BED1', end_color = 'B6BED1', fill_type = 'solid')
    wb.save('output_file.xlsx')
    wb.close()

    colourF = []
    for i in range(len(oneday_df)+2):
        if i % 2 == 0:
            if i != 0:
                colourF.append('C' + str(i))

    wb = load_workbook('output_file.xlsx')
    ws = wb['One Day Analyses']
    for i in range(len(colourF)):
        ws[(colourF[i])].fill = PatternFill(start_color = 'B6BED1', end_color = 'B6BED1', fill_type = 'solid')
    wb.save('output_file.xlsx')
    wb.close()

    colourG = []
    for i in range(len(oneday_df)+2):
        if i % 2 == 0:
            if i != 0:
                colourG.append('D' + str(i))

    wb = load_workbook('output_file.xlsx')
    ws = wb['One Day Analyses']
    for i in range(len(colourG)):
        ws[(colourG[i])].fill = PatternFill(start_color = 'B6BED1', end_color = 'B6BED1', fill_type = 'solid')
    wb.save('output_file.xlsx')
    wb.close()

    colourH = []
    for i in range(len(oneday_df)+2):
        if i % 2 == 0:
            if i != 0:
                colourH.append('E' + str(i))

    wb = load_workbook('output_file.xlsx')
    ws = wb['One Day Analyses']
    for i in range(len(colourH)):
        ws[(colourH[i])].fill = PatternFill(start_color = 'B6BED1', end_color = 'B6BED1', fill_type = 'solid')
    wb.save('output_file.xlsx')
    wb.close()

    colourI = []
    for i in range(len(oneday_df)+2):
        if i % 2 == 0:
            if i != 0:
                colourI.append('F' + str(i))

    wb = load_workbook('output_file.xlsx')
    ws = wb['One Day Analyses']
    for i in range(len(colourI)):
        ws[(colourI[i])].fill = PatternFill(start_color = 'B6BED1', end_color = 'B6BED1', fill_type = 'solid')
    wb.save('output_file.xlsx')
    wb.close()

    #3.4.3 One Day Probability Data Frame
    colourJ = []
    for i in range(len(ddf_df)+2):
        if i % 2 == 0:
            if i != 0:
                colourJ.append('A' + str(i))

    wb = load_workbook('output_file.xlsx')
    ws = wb['One Day Probability Data Frame']
    for i in range(len(colourJ)):
        ws[(colourJ[i])].fill = PatternFill(start_color = 'B6BED1', end_color = 'B6BED1', fill_type = 'solid')
    wb.save('output_file.xlsx')
    wb.close()

    colourK = []
    for i in range(len(ddf_df)+2):
        if i % 2 == 0:
            if i != 0:
                colourK.append('B' + str(i))

    wb = load_workbook('output_file.xlsx')
    ws = wb['One Day Probability Data Frame']
    for i in range(len(colourK)):
        ws[(colourK[i])].fill = PatternFill(start_color = 'B6BED1', end_color = 'B6BED1', fill_type = 'solid')
    wb.save('output_file.xlsx')
    wb.close()

    #3.4.4 Gumbel Estimate (mm_day)
    colourL = []
    for i in range(len(gumbelest_df)+2):
        if i % 2 == 0:
            if i != 0:
                colourL.append('A' + str(i))

    wb = load_workbook('output_file.xlsx')
    ws = wb['Gumbel Estimate (mm_day)']
    for i in range(len(colourL)):
        ws[(colourL[i])].fill = PatternFill(start_color = 'B6BED1', end_color = 'B6BED1', fill_type = 'solid')
    wb.save('output_file.xlsx')
    wb.close()

    colourM = []
    for i in range(len(gumbelest_df)+2):
        if i % 2 == 0:
            if i != 0:
                colourM.append('B' + str(i))

    wb = load_workbook('output_file.xlsx')
    ws = wb['Gumbel Estimate (mm_day)']
    for i in range(len(colourM)):
        ws[(colourM[i])].fill = PatternFill(start_color = 'B6BED1', end_color = 'B6BED1', fill_type = 'solid')
    wb.save('output_file.xlsx')
    wb.close()

    colourN = []
    for i in range(len(gumbelest_df)+2):
        if i % 2 == 0:
            if i != 0:
                colourN.append('C' + str(i))

    wb = load_workbook('output_file.xlsx')
    ws = wb['Gumbel Estimate (mm_day)']
    for i in range(len(colourN)):
        ws[(colourN[i])].fill = PatternFill(start_color = 'B6BED1', end_color = 'B6BED1', fill_type = 'solid')
    wb.save('output_file.xlsx')
    wb.close()

    colourO = []
    for i in range(len(gumbelest_df)+2):
        if i % 2 == 0:
            if i != 0:
                colourO.append('D' + str(i))

    wb = load_workbook('output_file.xlsx')
    ws = wb['Gumbel Estimate (mm_day)']
    for i in range(len(colourO)):
        ws[(colourO[i])].fill = PatternFill(start_color = 'B6BED1', end_color = 'B6BED1', fill_type = 'solid')
    wb.save('output_file.xlsx')
    wb.close()

    colourP = []
    for i in range(len(gumbelest_df)+2):
        if i % 2 == 0:
            if i != 0:
                colourP.append('E' + str(i))

    wb = load_workbook('output_file.xlsx')
    ws = wb['Gumbel Estimate (mm_day)']
    for i in range(len(colourP)):
        ws[(colourP[i])].fill = PatternFill(start_color = 'B6BED1', end_color = 'B6BED1', fill_type = 'solid')
    wb.save('output_file.xlsx')
    wb.close()

    colourQ = []
    for i in range(len(gumbelest_df)+2):
        if i % 2 == 0:
            if i != 0:
                colourQ.append('F' + str(i))

    wb = load_workbook('output_file.xlsx')
    ws = wb['Gumbel Estimate (mm_day)']
    for i in range(len(colourQ)):
        ws[(colourQ[i])].fill = PatternFill(start_color = 'B6BED1', end_color = 'B6BED1', fill_type = 'solid')
    wb.save('output_file.xlsx')
    wb.close()

    colourR = []
    for i in range(len(gumbelest_df)+2):
        if i % 2 == 0:
            if i != 0:
                colourR.append('G' + str(i))

    wb = load_workbook('output_file.xlsx')
    ws = wb['Gumbel Estimate (mm_day)']
    for i in range(len(colourR)):
        ws[(colourR[i])].fill = PatternFill(start_color = 'B6BED1', end_color = 'B6BED1', fill_type = 'solid')
    wb.save('output_file.xlsx')
    wb.close()

    colourS = []
    for i in range(len(gumbelest_df)+2):
        if i % 2 == 0:
            if i != 0:
                colourS.append('H' + str(i))

    wb = load_workbook('output_file.xlsx')
    ws = wb['Gumbel Estimate (mm_day)']
    for i in range(len(colourS)):
        ws[(colourS[i])].fill = PatternFill(start_color = 'B6BED1', end_color = 'B6BED1', fill_type = 'solid')
    wb.save('output_file.xlsx')
    wb.close()

    colourT = []
    for i in range(len(gumbelest_df)+2):
        if i % 2 == 0:
            if i != 0:
                colourT.append('I' + str(i))

    wb = load_workbook('output_file.xlsx')
    ws = wb['Gumbel Estimate (mm_day)']
    for i in range(len(colourT)):
        ws[(colourT[i])].fill = PatternFill(start_color = 'B6BED1', end_color = 'B6BED1', fill_type = 'solid')
    wb.save('output_file.xlsx')
    wb.close()

    #3.4.5 Gumbel Est and Full Series
    colourU = []
    for i in range(len(fullgumbel_df)+2):
        if i % 2 == 0:
            if i != 0:
                colourU.append('A' + str(i))

    wb = load_workbook('output_file.xlsx')
    ws = wb['Gumbel Est and Full Series']
    for i in range(len(colourU)):
        ws[(colourU[i])].fill = PatternFill(start_color = 'B6BED1', end_color = 'B6BED1', fill_type = 'solid')
    wb.save('output_file.xlsx')
    wb.close()

    colourV = []
    for i in range(len(fullgumbel_df)+2):
        if i % 2 == 0:
            if i != 0:
                colourV.append('B' + str(i))

    wb = load_workbook('output_file.xlsx')
    ws = wb['Gumbel Est and Full Series']
    for i in range(len(colourV)):
        ws[(colourV[i])].fill = PatternFill(start_color = 'B6BED1', end_color = 'B6BED1', fill_type = 'solid')
    wb.save('output_file.xlsx')
    wb.close()

    colourW = []
    for i in range(len(fullgumbel_df)+2):
        if i % 2 == 0:
            if i != 0:
                colourW.append('C' + str(i))

    wb = load_workbook('output_file.xlsx')
    ws = wb['Gumbel Est and Full Series']
    for i in range(len(colourW)):
        ws[(colourW[i])].fill = PatternFill(start_color = 'B6BED1', end_color = 'B6BED1', fill_type = 'solid')
    wb.save('output_file.xlsx')
    wb.close()

    colourX = []
    for i in range(len(fullgumbel_df)+2):
        if i % 2 == 0:
            if i != 0:
                colourX.append('D' + str(i))

    wb = load_workbook('output_file.xlsx')
    ws = wb['Gumbel Est and Full Series']
    for i in range(len(colourX)):
        ws[(colourX[i])].fill = PatternFill(start_color = 'B6BED1', end_color = 'B6BED1', fill_type = 'solid')
    wb.save('output_file.xlsx')
    wb.close()

    colourY = []
    for i in range(len(fullgumbel_df)+2):
        if i % 2 == 0:
            if i != 0:
                colourY.append('E' + str(i))

    wb = load_workbook('output_file.xlsx')
    ws = wb['Gumbel Est and Full Series']
    for i in range(len(colourY)):
        ws[(colourY[i])].fill = PatternFill(start_color = 'B6BED1', end_color = 'B6BED1', fill_type = 'solid')
    wb.save('output_file.xlsx')
    wb.close()

    #3.4.6 Residual Mass
    colourZ = []
    for i in range(len(resmass_df)+2):
        if i % 2 == 0:
            if i != 0:
                colourZ.append('A' + str(i))

    wb = load_workbook('output_file.xlsx')
    ws = wb['Residual Mass']
    for i in range(len(colourZ)):
        ws[(colourZ[i])].fill = PatternFill(start_color = 'B6BED1', end_color = 'B6BED1', fill_type = 'solid')
    wb.save('output_file.xlsx')
    wb.close()

    colourAA = []
    for i in range(len(resmass_df)+2):
        if i % 2 == 0:
            if i != 0:
                colourAA.append('B' + str(i))

    wb = load_workbook('output_file.xlsx')
    ws = wb['Residual Mass']
    for i in range(len(colourAA)):
        ws[(colourAA[i])].fill = PatternFill(start_color = 'B6BED1', end_color = 'B6BED1', fill_type = 'solid')
    wb.save('output_file.xlsx')
    wb.close()

    colourAB = []
    for i in range(len(resmass_df)+2):
        if i % 2 == 0:
            if i != 0:
                colourAB.append('C' + str(i))

    wb = load_workbook('output_file.xlsx')
    ws = wb['Residual Mass']
    for i in range(len(colourAB)):
        ws[(colourAB[i])].fill = PatternFill(start_color = 'B6BED1', end_color = 'B6BED1', fill_type = 'solid')
    wb.save('output_file.xlsx')
    wb.close()

    colourAC = []
    for i in range(len(resmass_df)+2):
        if i % 2 == 0:
            if i != 0:
                colourAC.append('D' + str(i))

    wb = load_workbook('output_file.xlsx')
    ws = wb['Residual Mass']
    for i in range(len(colourAC)):
        ws[(colourAC[i])].fill = PatternFill(start_color = 'B6BED1', end_color = 'B6BED1', fill_type = 'solid')
    wb.save('output_file.xlsx')
    wb.close()

    colourAD = []
    for i in range(len(resmass_df)+2):
        if i % 2 == 0:
            if i != 0:
                colourAD.append('E' + str(i))

    wb = load_workbook('output_file.xlsx')
    ws = wb['Residual Mass']
    for i in range(len(colourAD)):
        ws[(colourAD[i])].fill = PatternFill(start_color = 'B6BED1', end_color = 'B6BED1', fill_type = 'solid')
    wb.save('output_file.xlsx')
    wb.close()

    colourAE_extra = []
    for i in range(len(resmass_df)+2):
        if i % 2 == 0:
            if i != 0:
                colourAE_extra.append('F' + str(i))

    wb = load_workbook('output_file.xlsx')
    ws = wb['Residual Mass']
    for i in range(len(colourAE_extra)):
        ws[(colourAE_extra[i])].fill = PatternFill(start_color = 'B6BED1', end_color = 'B6BED1', fill_type = 'solid')
    wb.save('output_file.xlsx')
    wb.close()

    #3.4.7 Monthly Rainfall (Box_Whisker)
    colourAE = []
    for i in range(len(monthly_df)+2):
        if i % 2 == 0:
            if i != 0:
                colourAE.append('A' + str(i))

    wb = load_workbook('output_file.xlsx')
    ws = wb['Monthly Rainfall (Box_Whisker)']
    for i in range(len(colourAE)):
        ws[(colourAE[i])].fill = PatternFill(start_color = 'B6BED1', end_color = 'B6BED1', fill_type = 'solid')
    wb.save('output_file.xlsx')
    wb.close()

    colourAF = []
    for i in range(len(monthly_df)+2):
        if i % 2 == 0:
            if i != 0:
                colourAF.append('B' + str(i))

    wb = load_workbook('output_file.xlsx')
    ws = wb['Monthly Rainfall (Box_Whisker)']
    for i in range(len(colourAF)):
        ws[(colourAF[i])].fill = PatternFill(start_color = 'B6BED1', end_color = 'B6BED1', fill_type = 'solid')
    wb.save('output_file.xlsx')
    wb.close()

    colourAG = []
    for i in range(len(monthly_df)+2):
        if i % 2 == 0:
            if i != 0:
                colourAG.append('C' + str(i))

    wb = load_workbook('output_file.xlsx')
    ws = wb['Monthly Rainfall (Box_Whisker)']
    for i in range(len(colourAG)):
        ws[(colourAG[i])].fill = PatternFill(start_color = 'B6BED1', end_color = 'B6BED1', fill_type = 'solid')
    wb.save('output_file.xlsx')
    wb.close()

    colourAH = []
    for i in range(len(monthly_df)+2):
        if i % 2 == 0:
            if i != 0:
                colourAH.append('D' + str(i))

    wb = load_workbook('output_file.xlsx')
    ws = wb['Monthly Rainfall (Box_Whisker)']
    for i in range(len(colourAH)):
        ws[(colourAH[i])].fill = PatternFill(start_color = 'B6BED1', end_color = 'B6BED1', fill_type = 'solid')
    wb.save('output_file.xlsx')
    wb.close()

    colourAI = []
    for i in range(len(monthly_df)+2):
        if i % 2 == 0:
            if i != 0:
                colourAI.append('E' + str(i))

    wb = load_workbook('output_file.xlsx')
    ws = wb['Monthly Rainfall (Box_Whisker)']
    for i in range(len(colourAI)):
        ws[(colourAI[i])].fill = PatternFill(start_color = 'B6BED1', end_color = 'B6BED1', fill_type = 'solid')
    wb.save('output_file.xlsx')
    wb.close()

    colourAJ = []
    for i in range(len(monthly_df)+2):
        if i % 2 == 0:
            if i != 0:
                colourAJ.append('F' + str(i))

    wb = load_workbook('output_file.xlsx')
    ws = wb['Monthly Rainfall (Box_Whisker)']
    for i in range(len(colourAJ)):
        ws[(colourAJ[i])].fill = PatternFill(start_color = 'B6BED1', end_color = 'B6BED1', fill_type = 'solid')
    wb.save('output_file.xlsx')
    wb.close()

    colourAK = []
    for i in range(len(monthly_df)+2):
        if i % 2 == 0:
            if i != 0:
                colourAK.append('G' + str(i))

    wb = load_workbook('output_file.xlsx')
    ws = wb['Monthly Rainfall (Box_Whisker)']
    for i in range(len(colourAK)):
        ws[(colourAK[i])].fill = PatternFill(start_color = 'B6BED1', end_color = 'B6BED1', fill_type = 'solid')
    wb.save('output_file.xlsx')
    wb.close()

    colourAL = []
    for i in range(len(monthly_df)+2):
        if i % 2 == 0:
            if i != 0:
                colourAL.append('H' + str(i))

    wb = load_workbook('output_file.xlsx')
    ws = wb['Monthly Rainfall (Box_Whisker)']
    for i in range(len(colourAL)):
        ws[(colourAL[i])].fill = PatternFill(start_color = 'B6BED1', end_color = 'B6BED1', fill_type = 'solid')
    wb.save('output_file.xlsx')
    wb.close()

    colourAM = []
    for i in range(len(monthly_df)+2):
        if i % 2 == 0:
            if i != 0:
                colourAM.append('I' + str(i))

    wb = load_workbook('output_file.xlsx')
    ws = wb['Monthly Rainfall (Box_Whisker)']
    for i in range(len(colourAM)):
        ws[(colourAM[i])].fill = PatternFill(start_color = 'B6BED1', end_color = 'B6BED1', fill_type = 'solid')
    wb.save('output_file.xlsx')
    wb.close()

    colourAN = []
    for i in range(len(monthly_df)+2):
        if i % 2 == 0:
            if i != 0:
                colourAN.append('J' + str(i))

    wb = load_workbook('output_file.xlsx')
    ws = wb['Monthly Rainfall (Box_Whisker)']
    for i in range(len(colourAN)):
        ws[(colourAN[i])].fill = PatternFill(start_color = 'B6BED1', end_color = 'B6BED1', fill_type = 'solid')
    wb.save('output_file.xlsx')
    wb.close()

    colourAO = []
    for i in range(len(monthly_df)+2):
        if i % 2 == 0:
            if i != 0:
                colourAO.append('K' + str(i))

    wb = load_workbook('output_file.xlsx')
    ws = wb['Monthly Rainfall (Box_Whisker)']
    for i in range(len(colourAO)):
        ws[(colourAO[i])].fill = PatternFill(start_color = 'B6BED1', end_color = 'B6BED1', fill_type = 'solid')
    wb.save('output_file.xlsx')
    wb.close()

    colourAP = []
    for i in range(len(monthly_df)+2):
        if i % 2 == 0:
            if i != 0:
                colourAP.append('L' + str(i))

    wb = load_workbook('output_file.xlsx')
    ws = wb['Monthly Rainfall (Box_Whisker)']
    for i in range(len(colourAP)):
        ws[(colourAP[i])].fill = PatternFill(start_color = 'B6BED1', end_color = 'B6BED1', fill_type = 'solid')
    wb.save('output_file.xlsx')
    wb.close()

    colourAQ = []
    for i in range(len(monthly_df)+2):
        if i % 2 == 0:
            if i != 0:
                colourAQ.append('M' + str(i))

    wb = load_workbook('output_file.xlsx')
    ws = wb['Monthly Rainfall (Box_Whisker)']
    for i in range(len(colourAQ)):
        ws[(colourAQ[i])].fill = PatternFill(start_color = 'B6BED1', end_color = 'B6BED1', fill_type = 'solid')
    wb.save('output_file.xlsx')
    wb.close()

    colourAR = []
    for i in range(len(monthly_df)+2):
        if i % 2 == 0:
            if i != 0:
                colourAR.append('N' + str(i))

    wb = load_workbook('output_file.xlsx')
    ws = wb['Monthly Rainfall (Box_Whisker)']
    for i in range(len(colourAR)):
        ws[(colourAR[i])].fill = PatternFill(start_color = 'B6BED1', end_color = 'B6BED1', fill_type = 'solid')
    wb.save('output_file.xlsx')
    wb.close()

    #3.4.8 LP (III) Estimate
    colourAS = []
    for i in range(len(lpiii_table_df)+2):
        if i % 2 == 0:
            if i != 0:
                colourAS.append('A' + str(i))

    wb = load_workbook('output_file.xlsx')
    ws = wb['LP (III) Estimate']
    for i in range(len(colourAS)):
        ws[(colourAS[i])].fill = PatternFill(start_color = 'B6BED1', end_color = 'B6BED1', fill_type = 'solid')
    wb.save('output_file.xlsx')
    wb.close()

    colourAT = []
    for i in range(len(lpiii_table_df)+2):
        if i % 2 == 0:
            if i != 0:
                colourAT.append('B' + str(i))

    wb = load_workbook('output_file.xlsx')
    ws = wb['LP (III) Estimate']
    for i in range(len(colourAT)):
        ws[(colourAT[i])].fill = PatternFill(start_color = 'B6BED1', end_color = 'B6BED1', fill_type = 'solid')
    wb.save('output_file.xlsx')
    wb.close()

    colourAU = []
    for i in range(len(lpiii_table_df)+2):
        if i % 2 == 0:
            if i != 0:
                colourAU.append('C' + str(i))

    wb = load_workbook('output_file.xlsx')
    ws = wb['LP (III) Estimate']
    for i in range(len(colourAU)):
        ws[(colourAU[i])].fill = PatternFill(start_color = 'B6BED1', end_color = 'B6BED1', fill_type = 'solid')
    wb.save('output_file.xlsx')
    wb.close()

    colourAV = []
    for i in range(len(lpiii_table_df)+2):
        if i % 2 == 0:
            if i != 0:
                colourAV.append('D' + str(i))

    wb = load_workbook('output_file.xlsx')
    ws = wb['LP (III) Estimate']
    for i in range(len(colourAV)):
        ws[(colourAV[i])].fill = PatternFill(start_color = 'B6BED1', end_color = 'B6BED1', fill_type = 'solid')
    wb.save('output_file.xlsx')
    wb.close()

    colourAW = []
    for i in range(len(lpiii_table_df)+2):
        if i % 2 == 0:
            if i != 0:
                colourAW.append('E' + str(i))

    wb = load_workbook('output_file.xlsx')
    ws = wb['LP (III) Estimate']
    for i in range(len(colourAW)):
        ws[(colourAW[i])].fill = PatternFill(start_color = 'B6BED1', end_color = 'B6BED1', fill_type = 'solid')
    wb.save('output_file.xlsx')
    wb.close()

    colourAX = []
    for i in range(len(lpiii_table_df)+2):
        if i % 2 == 0:
            if i != 0:
                colourAX.append('F' + str(i))

    wb = load_workbook('output_file.xlsx')
    ws = wb['LP (III) Estimate']
    for i in range(len(colourAX)):
        ws[(colourAX[i])].fill = PatternFill(start_color = 'B6BED1', end_color = 'B6BED1', fill_type = 'solid')
    wb.save('output_file.xlsx')
    wb.close()

    colourAY = []
    for i in range(len(lpiii_table_df)+2):
        if i % 2 == 0:
            if i != 0:
                colourAY.append('G' + str(i))

    wb = load_workbook('output_file.xlsx')
    ws = wb['LP (III) Estimate']
    for i in range(len(colourAY)):
        ws[(colourAY[i])].fill = PatternFill(start_color = 'B6BED1', end_color = 'B6BED1', fill_type = 'solid')
    wb.save('output_file.xlsx')
    wb.close()

    colourAZ = []
    for i in range(len(lpiii_table_df)+2):
        if i % 2 == 0:
            if i != 0:
                colourAZ.append('H' + str(i))

    wb = load_workbook('output_file.xlsx')
    ws = wb['LP (III) Estimate']
    for i in range(len(colourAZ)):
        ws[(colourAZ[i])].fill = PatternFill(start_color = 'B6BED1', end_color = 'B6BED1', fill_type = 'solid')
    wb.save('output_file.xlsx')
    wb.close()

    colourAAA = []
    for i in range(len(lpiii_table_df)+2):
        if i % 2 == 0:
            if i != 0:
                colourAAA.append('I' + str(i))

    wb = load_workbook('output_file.xlsx')
    ws = wb['LP (III) Estimate']
    for i in range(len(colourAAA)):
        ws[(colourAAA[i])].fill = PatternFill(start_color = 'B6BED1', end_color = 'B6BED1', fill_type = 'solid')
    wb.save('output_file.xlsx')
    wb.close()

    colourAAB = []
    for i in range(len(lpiii_table_df)+2):
        if i % 2 == 0:
            if i != 0:
                colourAAB.append('J' + str(i))

    wb = load_workbook('output_file.xlsx')
    ws = wb['LP (III) Estimate']
    for i in range(len(colourAAB)):
        ws[(colourAAB[i])].fill = PatternFill(start_color = 'B6BED1', end_color = 'B6BED1', fill_type = 'solid')
    wb.save('output_file.xlsx')
    wb.close()

    colourAAC = []
    for i in range(len(lpiii_table_df)+2):
        if i % 2 == 0:
            if i != 0:
                colourAAC.append('K' + str(i))

    wb = load_workbook('output_file.xlsx')
    ws = wb['LP (III) Estimate']
    for i in range(len(colourAAC)):
        ws[(colourAAC[i])].fill = PatternFill(start_color = 'B6BED1', end_color = 'B6BED1', fill_type = 'solid')
    wb.save('output_file.xlsx')
    wb.close()

    #3.4.9 LP Est, GEV and MEV
    colourAAD = []
    for i in range(len(gevmev_df)+2):
        if i % 2 == 0:
            if i != 0:
                colourAAD.append('A' + str(i))

    wb = load_workbook('output_file.xlsx')
    ws = wb['LP Est, GEV and MEV']
    for i in range(len(colourAAD)):
        ws[(colourAAD[i])].fill = PatternFill(start_color = 'B6BED1', end_color = 'B6BED1', fill_type = 'solid')
    wb.save('output_file.xlsx')
    wb.close()

    colourAAE = []
    for i in range(len(gevmev_df)+2):
        if i % 2 == 0:
            if i != 0:
                colourAAE.append('B' + str(i))

    wb = load_workbook('output_file.xlsx')
    ws = wb['LP Est, GEV and MEV']
    for i in range(len(colourAAE)):
        ws[(colourAAE[i])].fill = PatternFill(start_color = 'B6BED1', end_color = 'B6BED1', fill_type = 'solid')
    wb.save('output_file.xlsx')
    wb.close()

    colourAAF = []
    for i in range(len(gevmev_df)+2):
        if i % 2 == 0:
            if i != 0:
                colourAAF.append('C' + str(i))

    wb = load_workbook('output_file.xlsx')
    ws = wb['LP Est, GEV and MEV']
    for i in range(len(colourAAF)):
        ws[(colourAAF[i])].fill = PatternFill(start_color = 'B6BED1', end_color = 'B6BED1', fill_type = 'solid')
    wb.save('output_file.xlsx')
    wb.close()

    colourAAG = []
    for i in range(len(gevmev_df)+2):
        if i % 2 == 0:
            if i != 0:
                colourAAG.append('D' + str(i))

    wb = load_workbook('output_file.xlsx')
    ws = wb['LP Est, GEV and MEV']
    for i in range(len(colourAAG)):
        ws[(colourAAG[i])].fill = PatternFill(start_color = 'B6BED1', end_color = 'B6BED1', fill_type = 'solid')
    wb.save('output_file.xlsx')
    wb.close()

    colourAAH = []
    for i in range(len(gevmev_df)+2):
        if i % 2 == 0:
            if i != 0:
                colourAAH.append('E' + str(i))

    wb = load_workbook('output_file.xlsx')
    ws = wb['LP Est, GEV and MEV']
    for i in range(len(colourAAH)):
        ws[(colourAAH[i])].fill = PatternFill(start_color = 'B6BED1', end_color = 'B6BED1', fill_type = 'solid')
    wb.save('output_file.xlsx')
    wb.close()

    colourAAI = []
    for i in range(len(gevmev_df)+2):
        if i % 2 == 0:
            if i != 0:
                colourAAI.append('F' + str(i))

    wb = load_workbook('output_file.xlsx')
    ws = wb['LP Est, GEV and MEV']
    for i in range(len(colourAAI)):
        ws[(colourAAI[i])].fill = PatternFill(start_color = 'B6BED1', end_color = 'B6BED1', fill_type = 'solid')
    wb.save('output_file.xlsx')
    wb.close()

    #3.5 Insert Excel Headers and Information Boxes (.png)
    wb = load_workbook('output_file.xlsx')
    ws = wb['Introduction']

    img = openpyxl.drawing.image.Image(python_path + 'introduction.png')
    img.width = 795
    img.height = 1125
    img.anchor = 'A1'
    ws.add_image(img)

    wb.save('output_file.xlsx')
    wb.close()

    wb = load_workbook('output_file.xlsx')
    ws = wb['Original Rainfall Data']

    img = openpyxl.drawing.image.Image(python_path + 'info_headers (1).png')
    img.width = 640
    img.height = 345
    img.anchor = 'K1'
    ws.add_image(img)

    wb.save('output_file.xlsx')
    wb.close()

    wb = load_workbook('output_file.xlsx')
    ws = wb['One Day Analyses']

    img = openpyxl.drawing.image.Image(python_path + 'info_headers (2).png')
    img.width = 640
    img.height = 345
    img.anchor = 'K1'
    ws.add_image(img)

    wb.save('output_file.xlsx')
    wb.close()

    wb = load_workbook('output_file.xlsx')
    ws = wb['One Day Probability Data Frame']

    img = openpyxl.drawing.image.Image(python_path + 'info_headers (3).png')
    img.width = 640
    img.height = 345
    img.anchor = 'G1'
    ws.add_image(img)

    wb.save('output_file.xlsx')
    wb.close()

    wb = load_workbook('output_file.xlsx')
    ws = wb['Gumbel Estimate (mm_day)']

    img = openpyxl.drawing.image.Image(python_path + 'info_headers (4).png')
    img.width = 640
    img.height = 345
    img.anchor = 'N1'
    ws.add_image(img)

    wb.save('output_file.xlsx')
    wb.close()

    wb = load_workbook('output_file.xlsx')
    ws = wb['Gumbel Est and Full Series']

    img = openpyxl.drawing.image.Image(python_path + 'info_headers (5).png')
    img.width = 640
    img.height = 345
    img.anchor = 'J1'
    ws.add_image(img)

    wb.save('output_file.xlsx')
    wb.close()

    wb = load_workbook('output_file.xlsx')
    ws = wb['Residual Mass']

    img = openpyxl.drawing.image.Image(python_path + 'info_headers (6).png')
    img.width = 640
    img.height = 345
    img.anchor = 'K1'
    ws.add_image(img)

    wb.save('output_file.xlsx')
    wb.close()

    wb = load_workbook('output_file.xlsx')
    ws = wb['Monthly Rainfall (Box_Whisker)']

    img = openpyxl.drawing.image.Image(python_path + 'info_headers (7).png')
    img.width = 640
    img.height = 345
    img.anchor = 'S1'
    ws.add_image(img)

    wb.save('output_file.xlsx')
    wb.close()

    wb = load_workbook('output_file.xlsx')
    ws = wb['LP (III) Estimate']

    img = openpyxl.drawing.image.Image(python_path + 'info_headers (8).png')
    img.width = 640
    img.height = 345
    img.anchor = 'P1'
    ws.add_image(img)

    wb.save('output_file.xlsx')
    wb.close()

    wb = load_workbook('output_file.xlsx')
    ws = wb['Log Pearson (III)']

    img = openpyxl.drawing.image.Image(python_path + 'lp_introduction.png')
    img.width = 795
    img.height = 560
    img.anchor = 'A1'
    ws.add_image(img)

    wb.save('output_file.xlsx')
    wb.close()

    wb = load_workbook('output_file.xlsx')
    ws = wb['Log Pearson (III)']

    img = openpyxl.drawing.image.Image(external_path + 'pearson_est_table.png')
    img.width = 690
    img.height = 475
    img.anchor = 'B49'
    ws.add_image(img)

    wb.save('output_file.xlsx')
    wb.close()

    wb = load_workbook('output_file.xlsx')
    ws = wb['Log Pearson (III)']

    img = openpyxl.drawing.image.Image(external_path + 'wt_value_table.png')
    img.width = 370
    img.height = 295
    img.anchor = 'G32'
    ws.add_image(img)

    wb.save('output_file.xlsx')
    wb.close()

    wb = load_workbook('output_file.xlsx')
    ws = wb['Log Pearson (III)']

    img = openpyxl.drawing.image.Image(python_path + 'pearson_wt_value.png')
    img.width = 750
    img.height = 1060
    img.anchor = 'N1'
    ws.add_image(img)

    wb.save('output_file.xlsx')
    wb.close()

    wb = load_workbook('output_file.xlsx')
    ws = wb['GEV and MEV']

    img = openpyxl.drawing.image.Image(python_path + 'gev_mev_introduction.png')
    img.width = 795
    img.height = 560
    img.anchor = 'A1'
    ws.add_image(img)

    wb.save('output_file.xlsx')
    wb.close()

    wb = load_workbook('output_file.xlsx')
    ws = wb['GEV and MEV']

    img = openpyxl.drawing.image.Image(external_path + 'mev_var_02.png')
    img.width = 260
    img.height = 55
    img.anchor = 'F31'
    ws.add_image(img)

    wb.save('output_file.xlsx')
    wb.close()

    wb = load_workbook('output_file.xlsx')
    ws = wb['LP Est, GEV and MEV']

    img = openpyxl.drawing.image.Image(python_path + 'info_headers (9).png')
    img.width = 640
    img.height = 345
    img.anchor = 'K1'
    ws.add_image(img)

    wb.save('output_file.xlsx')
    wb.close()

    #**************************************************************************************************
    #4. Add Graphs to Excel Spreadsheet
    #4.1 Add Graphs to 'South Frequency Analyses'
    a = len(df)+1

    wb = load_workbook('output_file.xlsx')
    ws = wb['Original Rainfall Data']

    #Data for plotting
    y_values = Reference(ws, min_col=2, max_col=2, min_row=2, max_row=a)
    x_values = Reference(ws, min_col=1, max_col=1, min_row=2, max_row=a)

    #Create object of BarChart class
    chart1 = BarChart()
    chart1.style = 3 #Style Modification
    chart1.add_data(y_values)
    chart1.set_categories(x_values)

    #Set the title of the chart
    chart1.title = 'Daily Rainfall (Class 5mm/day)'

    #Set the title of the x-axis
    chart1.x_axis.title = 'Date'

    #Set the title of the y-axis
    chart1.y_axis.title = 'Rainfall (Class 5mm/day)'

    #Change bar filling and line color
    s = chart1.series[0]
    s.graphicalProperties.line.solidFill = '3B3D76'
    s.graphicalProperties.solidFill = '3B3D76'

    #Change bar size
    chart1.height = 11  # default is 7.5
    chart1.width = 22  # default is 15

    #Delete Legend
    chart1.legend = None

    #The top-left corner of the chart
    ws.add_chart(chart1,"H20")

    wb.save('output_file.xlsx')
    wb.close()

    #**************************************************************************************************
    #4.2 Add Graphs to 'One Day Analyses'
    b = len(oneday_df)+1

    wb = load_workbook('output_file.xlsx')
    ws = wb['One Day Analyses']

    #Data for plotting
    y_values = Reference(ws, min_col=2, max_col=3, min_row=1, max_row=b)
    x_values = Reference(ws, min_col=1, max_col=1, min_row=2, max_row=b)

    #Create object of BarChart class
    chart2 = BarChart()
    chart2.style = 3 #Style Modification
    chart2.add_data(data=y_values,titles_from_data=True)
    chart2.set_categories(x_values)

    chart2.legend.title = series

    #Set the title of the chart
    chart2.title = 'Rainfall Occurrence (Class (5mm/day))'

    #Set the title of the x-axis
    chart2.x_axis.title = 'Class Intervall (mm)'

    #Set the title of the y-axis
    chart2.y_axis.title = 'Occurrences (Days)'

    #Change bar filling and line color
    s = chart2.series[0]
    s.graphicalProperties.line.solidFill = '3B3D76'
    s.graphicalProperties.solidFill = '3B3D76'

    #Change bar size
    chart2.height = 11  # default is 7.5
    chart2.width = 22  # default is 15

    #The top-left corner of the chart
    ws.add_chart(chart2,"H20")

    wb.save('output_file.xlsx')
    wb.close()

    wb = load_workbook('output_file.xlsx')
    ws = wb['One Day Analyses']

    #Data for plotting
    y_values = Reference(ws, min_col=5, max_col=6, min_row=1, max_row=b)
    x_values = Reference(ws, min_col=1, max_col=1, min_row=2, max_row=b)

    #Create object of LineChart class
    chart2_2 = LineChart()
    chart2_2.style = 3 #Style Modification
    chart2_2.add_data(data=y_values,titles_from_data=True)
    chart2_2.set_categories(x_values)

    chart2_2.legend.title = series

    #set the title of the chart
    chart2_2.title = 'Rainfall Probability'

    #Set the title of the x-axis
    chart2_2.x_axis.title = 'Class Intervall (mm)'

    #Change bar filling and line color
    s = chart2_2.series[0]
    s.graphicalProperties.line.solidFill = '3B3D76'
    s.graphicalProperties.solidFill = '3B3D76'

    #Change bar size
    chart2_2.height = 11  # default is 7.5
    chart2_2.width = 22  # default is 15

    #The top-left corner of the chart
    ws.add_chart(chart2_2,"H43")

    wb.save('output_file.xlsx')
    wb.close()

    #**************************************************************************************************
    #4.3 Add Graphs to 'One Day Probability Data Frame'
    c = len(ddf_df)+1

    wb = load_workbook('output_file.xlsx')
    ws = wb['One Day Probability Data Frame']

    #Data for plotting
    y_values = Reference(ws, min_col=2, max_col=2, min_row=2, max_row=c)
    x_values = Reference(ws, min_col=1, max_col=1, min_row=2, max_row=c)

    #Create object of LineChart class
    chart3 = LineChart()
    chart3.style = 3 #Style Modification
    chart3.add_data(y_values)
    chart3.set_categories(x_values)

    #Set the title of the chart
    chart3.title = 'Probability of Rainfall Event'

    #Set the title of the x-axis
    chart3.x_axis.title = 'Return Period (Year)'

    #Set the title of the y-axis
    chart3.y_axis.title = '1-Day P Sum (mm)'

    #Change bar filling and line color
    s = chart3.series[0]
    s.graphicalProperties.line.solidFill = '3B3D76'
    s.graphicalProperties.solidFill = '3B3D76'

    #Change bar size
    chart3.height = 11  # default is 7.5
    chart3.width = 22  # default is 15

    #Delete Legend
    chart3.legend = None

    #The top-left corner of the chart
    ws.add_chart(chart3,"D20")

    wb.save('output_file.xlsx')
    wb.close()

    #**************************************************************************************************
    #4.4 Add Graphs to 'Gumbel Estimate (mm_day)'
    d = len(gumbelest_df)+1

    wb = load_workbook('output_file.xlsx')
    ws = wb['Gumbel Estimate (mm_day)']

    #Data for plotting
    y_values = Reference(ws, min_col=2, max_col=2, min_row=2, max_row=d)
    x_values = Reference(ws, min_col=1, max_col=1, min_row=2, max_row=d)

    #Create object of BarChart class
    chart5 = BarChart()
    chart5.style = 3 #Style Modification
    chart5.add_data(y_values)
    chart5.set_categories(x_values)

    #Set the title of the chart
    chart5.title = 'Maximum One Day Rainfall(mm) in a Year'

    #Set the title of the x-axis
    chart5.x_axis.title = 'Year'

    #Set the title of the y-axis
    chart5.y_axis.title = 'Maximum Rainfall (mm/day)'

    #Change bar filling and line color
    s = chart5.series[0]
    s.graphicalProperties.line.solidFill = '3B3D76'
    s.graphicalProperties.solidFill = '3B3D76'

    #Change bar size
    chart5.height = 11  # default is 7.5
    chart5.width = 22  # default is 15

    #Delete Legend
    chart5.legend = None

    #The top-left corner of the chart
    ws.add_chart(chart5,"K20")

    wb.save('output_file.xlsx')
    wb.close()

    wb = load_workbook('output_file.xlsx')
    ws = wb['Gumbel Estimate (mm_day)']

    #Data for plotting
    y_values = Reference(ws, min_col=9, max_col=9, min_row=2, max_row=d)
    x_values = Reference(ws, min_col=1, max_col=1, min_row=2, max_row=d)

    #Create object of LineChart class
    chart5 = LineChart()
    chart5.style = 3 #Style Modification
    chart5.add_data(y_values)
    chart5.set_categories(x_values)

    #Set the title of the chart
    chart5.title = 'Gumbel Estimate (mm_day)'

    #Set the title of the x-axis
    chart5.x_axis.title = 'Year'

    #Set the title of the y-axis
    chart5.y_axis.title = 'Gumbel Estimate (mm_day)'

    #Change bar filling and line color
    s = chart5.series[0]
    s.graphicalProperties.line.solidFill = '3B3D76'
    s.graphicalProperties.solidFill = '3B3D76'

    #Change bar size
    chart5.height = 11  # default is 7.5
    chart5.width = 22  # default is 15

    #Delete Legend
    chart5.legend = None

    #The top-left corner of the chart
    ws.add_chart(chart5,"K43")

    wb.save('output_file.xlsx')
    wb.close()

    #**************************************************************************************************
    #4.5 Add Graphs to 'Gumbel Est and Full Series'
    e = len(fullgumbel_df)+1

    wb = load_workbook('output_file.xlsx')
    ws = wb['Gumbel Est and Full Series']

    #Data for plotting
    y_values = Reference(ws, min_col=4, max_col=5, min_row=1, max_row=e)
    x_values = Reference(ws, min_col=1, max_col=1, min_row=2, max_row=e)

    #Create object of LineChart class
    chart6 = LineChart()
    chart6.style = 3 #Style Modification
    chart6.add_data(data=y_values, titles_from_data=True)
    chart6.set_categories(x_values)

    #Set the title of the chart
    chart6.title = 'Gumbel Estimate (mm_day) vs. Full Series (mm_day)'

    #Set the title of the x-axis
    chart6.x_axis.title = 'Gumbel Estimate (mm_day) and Full Series (mm_day)'

    #Set the title of the y-axis
    chart6.y_axis.title = 'Return Period'

    #Change bar filling and line color
    s = chart6.series[0]
    s.graphicalProperties.line.solidFill = '3B3D76'
    s.graphicalProperties.solidFill = '3B3D76'

    #Change bar size
    chart6.height = 11  # default is 7.5
    chart6.width = 22  # default is 15

    #The top-left corner of the chart
    ws.add_chart(chart6,"G20")

    wb.save('output_file.xlsx')
    wb.close()

    #**************************************************************************************************
    #4.6 Add Graphs to 'Residual Mass'
    f = len(resmass_df)+1

    wb = load_workbook('output_file.xlsx')
    ws = wb['Residual Mass']

    #Data for plotting
    y_values = Reference(ws, min_col=2, max_col=2, min_row=2, max_row=f)
    x_values = Reference(ws, min_col=1, max_col=1, min_row=2, max_row=f)

    #Create object of BarChart class
    chart7_1 = BarChart()
    chart7_1.style = 3 #Style Modification
    chart7_1.add_data(y_values)
    chart7_1.set_categories(x_values)

    #Set the title of the chart
    chart7_1.title = 'Total Rainfall in Year'

    #Set the title of the x-axis
    chart7_1.x_axis.title = 'Year'

    #Set the title of the y-axis
    chart7_1.y_axis.title = 'Rainfall (mm)'

    #Change bar filling and line color
    s = chart7_1.series[0]
    s.graphicalProperties.line.solidFill = '3B3D76'
    s.graphicalProperties.solidFill = '3B3D76'

    #Change bar size
    chart7_1.height = 11  # default is 7.5
    chart7_1.width = 22  # default is 15

    #Delete Legend
    chart7_1.legend = None

    #The top-left corner of the chart
    ws.add_chart(chart7_1,"H20")

    wb.save('output_file.xlsx')
    wb.close()

    wb = load_workbook('output_file.xlsx')
    ws = wb['Residual Mass']

    #Data for plotting
    y_values = Reference(ws, min_col=3, max_col=3, min_row=2, max_row=f)
    x_values = Reference(ws, min_col=1, max_col=1, min_row=2, max_row=f)

    #Create object of LineChart class
    chart7_2 = LineChart()
    chart7_2.style = 3 #Style Modification
    chart7_2.add_data(y_values)
    chart7_2.set_categories(x_values)

    #Set the title of the chart
    chart7_2.title = 'Mass Plot'

    #Set the title of the x-axis
    chart7_2.x_axis.title = 'Year'

    #Set the title of the y-axis
    chart7_2.y_axis.title = 'Accumulated Rainfall (mm)'

    #Change bar filling and line color
    s = chart7_2.series[0]
    s.graphicalProperties.line.solidFill = '3B3D76'
    s.graphicalProperties.solidFill = '3B3D76'

    #Change bar size
    chart7_2.height = 11  # default is 7.5
    chart7_2.width = 22  # default is 15

    #Delete Legend
    chart7_2.legend = None

    #The top-left corner of the chart
    ws.add_chart(chart7_2,"H43")

    wb.save('output_file.xlsx')
    wb.close()

    wb = load_workbook('output_file.xlsx')
    ws = wb['Residual Mass']

    #Data for plotting
    y_values = Reference(ws, min_col=6, max_col=6, min_row=2, max_row=f)
    x_values = Reference(ws, min_col=1, max_col=1, min_row=2, max_row=f)

    #Create object of LineChart class
    chart7_3 = LineChart()
    chart7_3.style = 3 #Style Modification
    chart7_3.add_data(y_values)
    chart7_3.set_categories(x_values)

    #Set the title of the chart
    chart7_3.title = 'Residual Mass'

    #Set the title of the x-axis
    chart7_3.x_axis.title = 'Year'

    #Set the title of the y-axis
    chart7_3.y_axis.title = 'Residual Mass (mm) (Accumulated Rainfall in Year - Average Rainfall)'

    #Change bar filling and line color
    s = chart7_3.series[0]
    s.graphicalProperties.line.solidFill = '3B3D76'
    s.graphicalProperties.solidFill = '3B3D76'

    #Change bar size
    chart7_3.height = 11  # default is 7.5
    chart7_3.width = 22  # default is 15

    #Delete Legend
    chart7_3.legend = None

    #The top-left corner of the chart
    ws.add_chart(chart7_3,"H66")

    wb.save('output_file.xlsx')
    wb.close()

    #**************************************************************************************************
    #4.7 Add Graphs to 'Monthly Rainfall (Box_Whisker)'
    g = len(monthly_df)+1

    trace0 = go.Box(
        y = monthly_df['January'],
        name = 'January'
    )
    trace1 = go.Box(
        y = monthly_df['February'],
        name = 'February'
    )
    trace2 = go.Box(
        y = monthly_df['March'],
        name = 'March'
    )
    trace3 = go.Box(
        y = monthly_df['April'],
        name = 'April'
    )
    trace4 = go.Box(
        y = monthly_df['May'],
        name = 'May'
    )
    trace5 = go.Box(
        y = monthly_df['June'],
        name = 'June'
    )
    trace6 = go.Box(
        y = monthly_df['July'],
        name = 'July'
    )
    trace7 = go.Box(
        y = monthly_df['August'],
        name = 'August'
    )
    trace8 = go.Box(
        y = monthly_df['September'],
        name = 'September'
    )
    trace9 = go.Box(
        y = monthly_df['October'],
        name = 'October'
    )
    trace10 = go.Box(
        y = monthly_df['November'],
        name = 'November'
    )
    trace11 = go.Box(
        y = monthly_df['December'],
        name = 'December'
    )

    data = [trace0, trace1, trace2, trace3, trace4, trace5, trace6, trace7, trace8, trace9, trace10, trace11]
    layout = go.Layout(title = 'Box and Whisker: Monthly Rainfall (mm)')

    fig = go.Figure(data=data, layout=layout)
    pyo.plot(fig)

    img_bytes = fig.to_image(format="png", width=832, height=567, scale=2)
    Image(img_bytes)
    fig.write_image(python_path + 'box_whisker.png')

    time.sleep(10)

    wb = load_workbook('output_file.xlsx')
    ws = wb['Monthly Rainfall (Box_Whisker)']

    img = openpyxl.drawing.image.Image(python_path + 'box_whisker.png')
    img.width = 832
    img.height = 567
    img.anchor = 'P20'
    ws.add_image(img)

    wb.save('output_file.xlsx')
    wb.close()

    wb = load_workbook('output_file.xlsx')
    ws = wb['Monthly Rainfall (Box_Whisker)']

    #Data for plotting
    y_values = Reference(ws, min_col=2, max_col=13, min_row=1, max_row=g)
    x_values = Reference(ws, min_col=1, max_col=1, min_row=2, max_row=g)

    #Create object of BarChart class
    chart9 = BarChart()
    chart9.style = 3 #Style Modification
    chart9.add_data(data=y_values, titles_from_data=True)
    chart9.set_categories(x_values)

    #Set the title of the chart
    chart9.title = 'Total Monthly Rainfall (mm)'

    #Set the title of the x-axis
    chart9.x_axis.title = 'Year'

    #Set the title of the y-axis
    chart9.y_axis.title = 'Total Rainfall (mm) in Month'

    #Change bar filling and line color
    s = chart9.series[0]
    s.graphicalProperties.line.solidFill = '3B3D76'
    s.graphicalProperties.solidFill = '3B3D76'

    #Change bar size
    chart9.height = 11  # default is 7.5
    chart9.width = 22  # default is 15

    #The top-left corner of the chart
    ws.add_chart(chart9,"P50")

    wb.save('output_file.xlsx')
    wb.close()

    #**************************************************************************************************
    #4.8 Add Graphs to 'LP (III) Estimate'
    h = len(lpiii_table_df)+1

    wb = load_workbook('output_file.xlsx')
    ws = wb['LP (III) Estimate']

    #Data for plotting
    y_values_h = Reference(ws, min_col=11, max_col=11, min_row=1, max_row=h)
    x_values = Reference(ws, min_col=7, max_col=7, min_row=2, max_row=h)

    #Create object of LineChart class
    chart10 = LineChart()
    chart10.style = 3
    chart10.set_categories(x_values)

    #Add series data to the chart
    chart10.add_data(y_values_h, titles_from_data=True)

    #Set the title of the chart
    chart10.title = 'Log Pearson (III) Estimate'

    #Set the title of the x-axis
    chart10.x_axis.title = 'Return Period (T)'

    #Set the title of the y-axis
    chart10.y_axis.title = 'Precipitation (mm/d)'

    #Change bar filling and line color
    s = chart10.series[0]
    s.graphicalProperties.line.solidFill = '3B3D76'
    s.graphicalProperties.solidFill = '3B3D76'

    #Change line size
    chart10.height = 11
    chart10.width = 22

    #Add the chart to the worksheet
    ws.add_chart(chart10, "M20")

    #Save and close the workbook
    wb.save('output_file.xlsx')
    wb.close()

    #**************************************************************************************************
    #4.9 Add Graphs to 'LP Est, GEV and MEV'
    i = len(gevmev_df)+1

    wb = load_workbook('output_file.xlsx')
    ws = wb['LP Est, GEV and MEV']

    #Data for plotting
    y_values = Reference(ws, min_col=2, max_col=6, min_row=1, max_row=i)
    x_values = Reference(ws, min_col=9, max_col=9, min_row=2, max_row=i)

    #Create object of LineChart class
    chart11 = LineChart()
    chart11.style = 3
    chart11.set_categories(x_values)

    #Add series data to the chart
    chart11.add_data(y_values, titles_from_data=True)

    #Set the title of the chart
    chart11.title = 'Full Series (mm/d) vs. Gumbel Estimate (mm/d) vs. Log Pearson (III) Estimate (mm/d) vs. GEV Estimate (mm/d) vs. MEV Estimate (mm/d)'

    #Set the title of the x-axis
    chart11.x_axis.title = 'Return Period (T)'

    #Set the title of the y-axis
    chart11.y_axis.title = 'Precipitation (mm/d)'

    #Change bar filling and line color
    s = chart11.series[0]
    s.graphicalProperties.line.solidFill = '3B3D76'
    s.graphicalProperties.solidFill = '3B3D76'# Change bar filling and line color

    #Change line size
    chart11.height = 11
    chart11.width = 22

    #Add the chart to the worksheet
    ws.add_chart(chart11, "H20")

    #Save and close the workbook
    wb.save('output_file.xlsx')
    wb.close()

    ###################################################################################################
    ################################ 5. Save Excel to Downloads Folder ################################
    ###################################################################################################
    #5.1 Copy Excel Spreadsheet to Downloads Folder
    tod = datetime.now()
    date = str(tod.year) + '-' + str(tod.month) + '-' + str(tod.day) + ' ' + str(tod.hour) + '-' + str(tod.minute) + '-' + str(tod.second)

    path = os.path.join(downloads_path, date + ' rainfall_data')
    os.mkdir(path)
    shutil.move(external_path + 'temp-plot.html', downloads_path + date + ' rainfall_data/')
    os.rename(downloads_path + date + ' rainfall_data/' + 'temp-plot.html', downloads_path + date + ' rainfall_data/' + 'interactive_boxandwhisker.html')

    wb = load_workbook('output_file.xlsx')
    wb.save(downloads_path + date + ' rainfall_data/' + date + ' rainfall_data.xlsx')
    wb.close()

    #**************************************************************************************************
    #5.2 Delete output_file.xlsx and box_whisker.png
    os.remove(external_path + 'output_file.xlsx')
    os.remove(python_path + 'box_whisker.png')
    os.remove(external_path + 'pearson_est_table.png')
    os.remove(external_path + 'wt_value_table.png')
    os.remove(external_path + 'mev_var_01.png')
    os.remove(external_path + 'mev_var_02.png')
    os.remove(external_path + 'mev_var_03.png')

    ###################################################################################################
    ###################################### 6. Display Message Box #####################################
    ###################################################################################################
    messagebox.showinfo("Report Completed", "An Excel Spreadsheet has been created and saved in your Output Folder. Thank you for using the Rainfall Data Analysis Program!")
    time.sleep(10)
    close_app

###################################################################################################
##################################### Graphical User Interface ####################################
###################################################################################################
def select_excel_file():
    file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
    if file_path:
        orig_rain_data_path_value.set(file_path)

def select_folder(var):
    folder_path = filedialog.askdirectory()
    if folder_path:
        var.set(folder_path)

def rainfall_analysis_user_input():
    #Retrieve the values from the StringVar variables
    orig_rain_data_path = orig_rain_data_path_value.get()
    external_path = external_path_value.get() + '/'
    python_path = python_path_value.get() + '/'
    downloads_path = downloads_path_value.get() + '/'

    messagebox.showinfo("Info", "Inputs have been captured for analysis and program will run.")
    #Call the continue_rainfall_analysis function
    continue_rainfall_analysis(orig_rain_data_path, external_path, python_path, downloads_path)

def close_app():
    root.destroy()

#Initialize the main window
root = tk.Tk()
root.title("Rainfall Data Analysis Program")

#Get the current working directory and set the icon path
icon_path = os.path.join('background_data/FA Logo.ico')
root.iconbitmap(icon_path)  # Set the custom icon

bold_font = tkfont.Font(root, size=10, weight="bold")

#Define StringVar variables to hold file paths
orig_rain_data_path_value = tk.StringVar()
external_path_value = tk.StringVar()
python_path_value = tk.StringVar()
downloads_path_value = tk.StringVar()

#Add welcome text
welcome_text = """WELCOME TO THE RAINFALL DATA ANALYSIS PROGRAM
*** PLEASE DO NOT CLOSE THIS WINDOW, UNTIL 'REPORT COMPLETED' MESSAGE BOX APPEARS.
(Please Scroll Down to see further Instructions)

This program is designed to take your daily rainfall data and generate a comprehensive report that provides detailed insights into various aspects of the data. The report includes:

1. Original Rainfall Data: Graphically represented by daily rainfall classified in 5mm/d increments.
2. One Day Analysis: Shows occurrence, exceedance, rate, and probability, along with Rainfall Occurrence and Probability graphs.
3. One Day Probability Data Frame: Presents return period and corresponding one-day probability, accompanied by a Probability of Rainfall Event graph.
4. Gumbel Estimate: Details the maximum one-day rainfall, observations, rank by maximum one-day rainfall, exceedance, non-exceedance, return period, return variables, and the Gumbel Estimate, with graphs illustrating maximum one-day rainfall in a year and the Gumbel Estimate.
5. Gumbel Estimate and Full Series: Includes return period, return variable, Gumbel Estimate, and Full Series data, with a graph comparing Gumbel Estimate (mm/day) vs. Full Series (mm/day).
6. Residual Mass: Displays yearly total rainfall, accumulated rainfall, average yearly rainfall, total rainfall minus average rainfall, and accumulated rainfall minus average rainfall, with accompanying graphs.
7. Monthly Rainfall (Box and Whisker): Lists monthly rainfall data and total yearly rainfall (mm), supplemented by a box and whisker graph and a total monthly rainfall graph.
8. Log Pearson (III): Presents frequency factors for gamma and Log-Pearson Type III distributions, return period, Wt value, log Qt, Pearson Estimate (mm/day), and Full Series (mm/day).
9. GEV and MEV: Provides csi, sigma, mu, and starting guesses for MEV (x0).
10. LP (III) Estimate: Includes observed max rainfall (mm/day), rank, log observed max, statistical calculations, return period, exceedance probability, non-exceedance probability, Wt, and LP (III) Estimate (mm/day), along with a Log Pearson (III) Estimate graph.
11. LP Est, GEV, and MEV: Compares return period, Full Series (mm/day), Gumbel Estimate (mm/day), Log Pearson (III) Estimate (mm/day), GEV Estimate (mm/day), and MEV Estimate (mm/day) with a comparative graph of all estimates.

"""

instructions_text = """INSTRUCTIONS FOR PREPARING THE RAINFALL DATA SET

To use this program, you will need to provide a daily rainfall data set in Excel format (.xlsx). Please ensure your data adheres to the following specifications:

1. Format: The file should be in Excel format with the extension .xlsx.
2. Daily Rainfall: The data should be recorded on a daily basis.
3. Header Row: The first row should contain the following EXACT labels:
   - A1: "Date"
   - B1: "Rainfall (mm/d)"
   - C1: "Patched/Recorded"
4. Data Columns: Each column should be formatted as follows:
   - Column A: Should be in DATE format (e.g., 14/03/2012).
   - Column B: Should be in NUMBER format (e.g., 0.02).
   - Column C: Should contain the text "Patched" or "Recorded" only. No other values are permitted.
5. Worksheet: Ensure the Excel spreadsheet contains only one worksheet, which can be named anything.
6. File Name: The Excel file can be named anything.
"""

#Create a frame for the text widget and scrollbar
text_frame = tk.Frame(root)
text_frame.pack(pady=10, padx=10, fill=tk.BOTH, expand=True)

#Create a scrollbar
scrollbar = tk.Scrollbar(text_frame)
scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

#Create a text widget to display the information
info_text = tk.Text(text_frame, wrap=tk.WORD, font=("Helvetica", 10), yscrollcommand=scrollbar.set)
info_text.insert(tk.END, welcome_text)
info_text.tag_add("welcome", "1.0", "1.end")
info_text.tag_config("welcome", font=bold_font)
info_text.insert(tk.END, instructions_text)
info_text.tag_add("instructions", "2.0", "2.end")
info_text.tag_config("instructions", font=bold_font)
info_text.pack(pady=10, padx=10, fill=tk.BOTH, expand=True)

#Configure the scrollbar
scrollbar.config(command=info_text.yview)

#User input section
input_frame = tk.Frame(root)
input_frame.pack(pady=10, padx=10)

tk.Label(input_frame, text="USER INPUT", font=bold_font).grid(row=0, columnspan=3, pady=5)

tk.Label(input_frame, text="Select Excel File with the Original Rainfall Data Set").grid(row=1, column=0, sticky=tk.W)
tk.Entry(input_frame, textvariable=orig_rain_data_path_value, width=50).grid(row=1, column=1, padx=5)
tk.Button(input_frame, text="Browse", command=select_excel_file).grid(row=1, column=2)

tk.Label(input_frame, text="Select the Folder that contains 'fa_patched_rainfall.exe'").grid(row=2, column=0, sticky=tk.W)
tk.Entry(input_frame, textvariable=external_path_value, width=50).grid(row=2, column=1, padx=5)
tk.Button(input_frame, text="Browse", command=lambda: select_folder(external_path_value)).grid(row=2, column=2)

tk.Label(input_frame, text="Select 'background_data' Folder").grid(row=3, column=0, sticky=tk.W)
tk.Entry(input_frame, textvariable=python_path_value, width=50).grid(row=3, column=1, padx=5)
tk.Button(input_frame, text="Browse", command=lambda: select_folder(python_path_value)).grid(row=3, column=2)

tk.Label(input_frame, text="Select Output Folder").grid(row=4, column=0, sticky=tk.W)
tk.Entry(input_frame, textvariable=downloads_path_value, width=50).grid(row=4, column=1, padx=5)
tk.Button(input_frame, text="Browse", command=lambda: select_folder(downloads_path_value)).grid(row=4, column=2)

#Add important note
important_note_text = """IMPORTANT TO NOTE
Ensuring the data is in the correct format is crucial for the program to function correctly and produce accurate results.
*** PLEASE DO NOT CLOSE THIS WINDOW, UNTIL 'REPORT COMPLETED' MESSAGE BOX APPEARS.
Thank you for using the Rainfall Data Analysis!
"""

important_note_label = tk.Label(root, text=important_note_text, wraplength=600, justify=tk.LEFT, font=("Helvetica", 10))
important_note_label.pack(pady=10)

#Add OK and Close buttons
button_frame = tk.Frame(root)
button_frame.pack(pady=10)

ok_button = tk.Button(button_frame, text="OK", command=rainfall_analysis_user_input)
ok_button.grid(row=0, column=0, padx=5)

close_button = tk.Button(button_frame, text="Close", command=close_app)
close_button.grid(row=0, column=1, padx=5)

#Run the main loop
root.mainloop()

#**************************************************************************************************
#**************************************************************************************************
